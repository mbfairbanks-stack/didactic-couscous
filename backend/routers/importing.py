"""File import/export: CSV parsing, xlsx import/export, PDF statement parsing."""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Request
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy.orm import Session
from sqlalchemy import select, func, distinct, text, nullslast
from sqlalchemy.exc import IntegrityError
from pydantic import BaseModel
from typing import Optional, List
import datetime
import tempfile, os, io, json, csv, re, math
from collections import defaultdict

import models
from database import get_db
from importer import import_xlsx
from category_utils import ensure_category

router = APIRouter()

# ---------------------------------------------------------------------------
# Smart CSV paste import
# ---------------------------------------------------------------------------

class ParseCsvRequest(BaseModel):
    text: str
    format: str = "auto"  # "auto", "amex", "td", "rbc", "cibc"


def _parse_amount(s: str) -> Optional[float]:
    if not s:
        return None
    cleaned = re.sub(r'[$,\s]', '', str(s))
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_date(s: str) -> Optional[str]:
    from datetime import datetime
    for fmt in ('%m/%d/%Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%y', '%Y/%m/%d',
                '%b %d, %Y', '%B %d, %Y', '%b %d %Y', '%B %d %Y',
                '%d-%b-%Y', '%d %b %Y', '%Y%m%d',
                '%d %b %y', '%d %B %y'):  # e.g. "19 May 26"
        try:
            return datetime.strptime(s.strip(), fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def _lookup_category(merchant: str, db: Session) -> Optional[str]:
    """Find category: check merchant_rules first, then transaction history."""
    # 1. Exact match in merchant_rules
    rule = db.execute(
        select(models.MerchantRule).where(models.MerchantRule.merchant_pattern == merchant)
    ).scalar_one_or_none()
    if rule:
        return rule.category

    # 2. Partial match in merchant_rules
    rules = db.execute(select(models.MerchantRule)).scalars().all()
    merchant_lower = merchant.lower()
    for r in rules:
        if r.merchant_pattern.lower() in merchant_lower or merchant_lower in r.merchant_pattern.lower():
            return r.category

    # 3. Exact match in transaction history
    rows = db.execute(
        select(models.Transaction.category, func.count(models.Transaction.id).label("cnt"))
        .where(models.Transaction.merchant == merchant)
        .group_by(models.Transaction.category)
        .order_by(func.count(models.Transaction.id).desc())
    ).all()
    if rows:
        return rows[0].category

    # 4. Fuzzy match in transaction history
    words = [w for w in re.sub(r'[^a-z\s]', '', merchant.lower()).split() if len(w) > 3]
    if not words:
        return None
    search_word = words[0]
    rows = db.execute(
        select(models.Transaction.category, func.count(models.Transaction.id).label("cnt"))
        .where(func.lower(models.Transaction.merchant).contains(search_word))
        .group_by(models.Transaction.category)
        .order_by(func.count(models.Transaction.id).desc())
    ).all()
    return rows[0].category if rows else None


@router.post("/parse-csv")
def parse_csv(body: ParseCsvRequest, db: Session = Depends(get_db)):
    """Parse CSV from AMEX/Visa/MC and return rows with suggested categories.
    Handles both headered (comma) and headerless tab-delimited (TD) formats.
    """
    raw = body.text.strip()
    if not raw:
        return {"rows": []}

    lines = [l.strip() for l in raw.splitlines() if l.strip()]

    # Skip a header row if present (e.g. "Date  Description  Amount")
    if lines and re.search(r'\bdate\b', lines[0], re.IGNORECASE) and re.search(r'\bdescription\b', lines[0], re.IGNORECASE):
        lines = lines[1:]

    # ── Detect AMEX multi-line paste format ──────────────────────────────────
    # Pattern: groups of 3 lines — date, merchant, $amount (repeating)
    # e.g.:
    #   19 May 26
    #   COUNTRY PAWS DOGPLEX 00 LONDON
    #   $6.10
    def _is_amex_multiline(lines: list) -> bool:
        if len(lines) < 3:
            return False
        # Check first few groups — allow remainder lines to be off
        matches = 0
        for i in range(0, min(len(lines) - 2, 12), 3):
            if _parse_date(lines[i]) and re.match(r'^-?\$?[\d,]+\.?\d*$', lines[i + 2]):
                matches += 1
        return matches >= 1

    if _is_amex_multiline(lines):
        parsed = []
        for i in range(0, len(lines) - 2, 3):
            date_str = lines[i].strip()
            merchant = lines[i + 1].strip()
            amt_str = lines[i + 2].strip()
            parsed_date = _parse_date(date_str)
            amount = _parse_amount(amt_str)
            if not parsed_date or not merchant or not amount or amount <= 0:
                continue
            suggested = _lookup_category(merchant, db)
            # Check if already in transactions
            is_dup = db.execute(
                select(models.Transaction).where(
                    models.Transaction.date == datetime.date.fromisoformat(parsed_date),
                    models.Transaction.merchant == merchant,
                    func.round(models.Transaction.amount, 2) == round(amount, 2),
                )
            ).first() is not None
            parsed.append({
                "date": parsed_date,
                "merchant": merchant,
                "amount": round(amount, 2),
                "suggested_category": suggested or "",
                "confidence": "high" if suggested else "low",
                "is_duplicate": is_dup,
            })
        return {"rows": parsed}

    # Detect headerless format: tab or comma delimited, first cell is a date
    # TD format: date,merchant,debit,credit,balance  (or tab-separated)
    def _split_first(line: str):
        if '\t' in line:
            return [p.strip() for p in line.split('\t')]
        return next(csv.reader([line]))

    def _is_headerless(line: str) -> bool:
        parts = _split_first(line)
        return len(parts) >= 3 and bool(_parse_date(parts[0]))

    use_td = bool(lines) and _is_headerless(lines[0])

    parsed = []

    if use_td:
        for line in lines:
            if not line.strip():
                continue
            parts = _split_first(line)
            if len(parts) < 3:
                continue
            date_str = parts[0]
            merchant = parts[1]
            # col 2 = debit (expense), col 3 = credit (payment), col 4 = balance
            debit = _parse_amount(parts[2]) if len(parts) > 2 else None
            credit = _parse_amount(parts[3]) if len(parts) > 3 else None

            parsed_date = _parse_date(date_str)
            if not parsed_date or not merchant:
                continue

            # Only import debits (expenses); skip credits/payments
            if debit and debit > 0:
                amount = debit
            else:
                continue

            suggested = _lookup_category(merchant, db)
            # Check if already in transactions
            is_dup = db.execute(
                select(models.Transaction).where(
                    models.Transaction.date == datetime.date.fromisoformat(parsed_date),
                    models.Transaction.merchant == merchant,
                    func.round(models.Transaction.amount, 2) == round(amount, 2),
                )
            ).first() is not None
            parsed.append({
                "date": parsed_date,
                "merchant": merchant,
                "amount": round(amount, 2),
                "suggested_category": suggested or "",
                "confidence": "high" if suggested else "low",
                "is_duplicate": is_dup,
            })
    else:
        # Header-based CSV (AMEX, RBC, CIBC, etc.)
        reader = csv.reader(io.StringIO(raw))
        rows = list(reader)
        if len(rows) < 2:
            return {"rows": []}

        header = [h.strip().lower().replace(' ', '_') for h in rows[0]]

        def col(names):
            for n in names:
                for i, h in enumerate(header):
                    if n in h:
                        return i
            return None

        date_col  = col(['date', 'transaction_date'])
        desc_col  = col(['description', 'desc', 'merchant', 'name', 'payee'])
        amt_col   = col(['amount'])
        debit_col = col(['debit'])
        cad_col   = col(['cad$', 'cad'])

        for row in rows[1:]:
            if not row or all(not c.strip() for c in row):
                continue
            merchant = row[desc_col].strip() if desc_col is not None and desc_col < len(row) else ""
            date_str = row[date_col].strip() if date_col is not None and date_col < len(row) else ""
            if not merchant or not date_str:
                continue

            amount = None
            if amt_col is not None and amt_col < len(row):
                amount = _parse_amount(row[amt_col])
            if (amount is None or amount <= 0) and debit_col is not None and debit_col < len(row):
                amount = _parse_amount(row[debit_col])
            if (amount is None or amount <= 0) and cad_col is not None and cad_col < len(row):
                amount = _parse_amount(row[cad_col])

            if amount is None or amount <= 0:
                continue

            parsed_date = _parse_date(date_str)
            if not parsed_date:
                continue

            suggested = _lookup_category(merchant, db)
            # Check if already in transactions
            is_dup = db.execute(
                select(models.Transaction).where(
                    models.Transaction.date == datetime.date.fromisoformat(parsed_date),
                    models.Transaction.merchant == merchant,
                    func.round(models.Transaction.amount, 2) == round(amount, 2),
                )
            ).first() is not None
            parsed.append({
                "date": parsed_date,
                "merchant": merchant,
                "amount": round(amount, 2),
                "suggested_category": suggested or "",
                "confidence": "high" if suggested else "low",
                "is_duplicate": is_dup,
            })

    return {"rows": parsed}


class CsvImportRow(BaseModel):
    date: str
    merchant: str
    amount: float
    category: str
    source: Optional[str] = None


@router.post("/import-csv-rows")
def import_csv_rows(rows: List[CsvImportRow], db: Session = Depends(get_db)):
    """Import reviewed CSV rows into the transactions table, skipping exact duplicates."""
    from datetime import datetime
    imported = 0
    skipped_duplicates = 0
    for row in rows:
        try:
            d = datetime.strptime(row.date, '%Y-%m-%d').date()
        except ValueError:
            continue
        dup = db.execute(
            select(models.Transaction).where(
                models.Transaction.date == d,
                models.Transaction.merchant == row.merchant,
                func.round(models.Transaction.amount, 2) == round(row.amount, 2),
            )
        ).first()
        if dup:
            skipped_duplicates += 1
            continue
        source = getattr(row, 'source', None) or "csv_import"
        row.category = ensure_category(db, row.category)
        txn = models.Transaction(
            date=d,
            merchant=row.merchant,
            amount=row.amount,
            category=row.category,
            year=d.year,
            month=d.month,
            source=source,
        )
        db.add(txn)
        imported += 1
    db.commit()
    return {"imported": imported, "skipped_duplicates": skipped_duplicates}


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------

@router.post("/import")
async def import_file(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files are supported")

    content = await file.read()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        counts = import_xlsx(tmp_path, db)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Import failed: {type(e).__name__}: {e}")
    finally:
        os.unlink(tmp_path)

    return {
        "message": "Import complete",
        "transactions_imported": counts["transactions"],
        "income_imported": counts["income"],
        "records_updated": counts["skipped"],
    }


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

@router.get("/export")
def export_xlsx(year: int, month: Optional[int] = None, db: Session = Depends(get_db)):
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # Transactions sheet
    ws_txn = wb.active
    ws_txn.title = "Transactions"
    headers = ["Date", "Merchant", "Amount", "Category", "Month", "Year", "Source"]
    for col, h in enumerate(headers, 1):
        cell = ws_txn.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.font = Font(bold=True, color="FFFFFF")

    q = select(models.Transaction).where(models.Transaction.year == year)
    if month:
        q = q.where(models.Transaction.month == month)
    q = q.order_by(models.Transaction.date)
    txns = db.execute(q).scalars().all()

    for row_idx, txn in enumerate(txns, 2):
        ws_txn.cell(row=row_idx, column=1, value=txn.date)
        ws_txn.cell(row=row_idx, column=2, value=txn.merchant)
        ws_txn.cell(row=row_idx, column=3, value=txn.amount)
        ws_txn.cell(row=row_idx, column=4, value=txn.category)
        ws_txn.cell(row=row_idx, column=5, value=txn.month)
        ws_txn.cell(row=row_idx, column=6, value=txn.year)
        ws_txn.cell(row=row_idx, column=7, value=txn.source)

    # Monthly Summary sheet
    ws_sum = wb.create_sheet("Monthly Summary")
    sum_headers = ["Month", "Income", "Expenses", "Balance"]
    for col, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    monthly = monthly_summary(year, db)
    for row_idx, m in enumerate(monthly, 2):
        month_names = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
                       "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        ws_sum.cell(row=row_idx, column=1, value=month_names[m["month"]])
        ws_sum.cell(row=row_idx, column=2, value=m["income"])
        ws_sum.cell(row=row_idx, column=3, value=m["expenses"])
        ws_sum.cell(row=row_idx, column=4, value=m["balance"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"budget_{year}.xlsx" if not month else f"budget_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── PDF Bank Statement Parser ────────────────────────────────────────────────

@router.post("/parse-pdf")
async def parse_pdf(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Parse a bank statement PDF and extract transactions."""
    try:
        import pdfplumber
    except ImportError:
        raise HTTPException(500, "pdfplumber not installed")

    contents = await file.read()
    rows_out = []
    page_count = 0

    try:
        with pdfplumber.open(io.BytesIO(contents)) as pdf:
            page_count = len(pdf.pages)
            for page in pdf.pages:
                # Try table extraction first
                tables = page.extract_tables()
                for table in tables:
                    if not table or len(table) < 2:
                        continue
                    # Find header row
                    header = [str(c).lower().strip() if c else "" for c in table[0]]

                    def col_idx(names):
                        for name in names:
                            for i, h in enumerate(header):
                                if name in h:
                                    return i
                        return None

                    date_col = col_idx(["date", "transaction date", "posted"])
                    desc_col = col_idx(["description", "desc", "merchant", "details", "activity", "payee", "transaction"])
                    debit_col = col_idx(["debit", "withdrawal", "charges", "amount"])
                    credit_col = col_idx(["credit", "payment", "deposit"])
                    amt_col = col_idx(["amount"]) if debit_col is None else None

                    if date_col is None or desc_col is None:
                        continue

                    for row in table[1:]:
                        if not row or all(not c for c in row):
                            continue
                        date_str = str(row[date_col]).strip() if date_col < len(row) and row[date_col] else ""
                        merchant = str(row[desc_col]).strip() if desc_col < len(row) and row[desc_col] else ""
                        if not date_str or not merchant or merchant.lower() in ("", "none"):
                            continue

                        parsed_date = _parse_date(date_str)
                        if not parsed_date:
                            continue

                        amount = None
                        # Try debit column first
                        if debit_col is not None and debit_col < len(row) and row[debit_col]:
                            amount = _parse_amount(str(row[debit_col]))
                        # Fall back to amount column
                        if (amount is None or amount <= 0) and amt_col is not None and amt_col < len(row) and row[amt_col]:
                            val = _parse_amount(str(row[amt_col]))
                            if val and val > 0:
                                amount = val

                        if amount is None or amount <= 0:
                            continue

                        # Skip payment/credit rows
                        if credit_col is not None and credit_col < len(row) and row[credit_col]:
                            credit_val = _parse_amount(str(row[credit_col]))
                            if credit_val and credit_val > 0 and (amount is None or amount == 0):
                                continue

                        is_dup = db.execute(
                            select(models.Transaction).where(
                                models.Transaction.date == datetime.date.fromisoformat(parsed_date),
                                models.Transaction.merchant == merchant,
                                func.round(models.Transaction.amount, 2) == round(amount, 2),
                            )
                        ).first() is not None

                        suggested = _lookup_category(merchant, db)
                        rows_out.append({
                            "date": parsed_date,
                            "merchant": merchant,
                            "amount": round(amount, 2),
                            "suggested_category": suggested or "",
                            "confidence": "high" if suggested else "low",
                            "is_duplicate": is_dup,
                        })

                # If no tables found, try text-based extraction
                if not rows_out:
                    text = page.extract_text() or ""
                    lines = [l.strip() for l in text.splitlines() if l.strip()]
                    # Look for lines matching: date + description + amount pattern
                    for line in lines:
                        # Try to match: date at start, amount at end
                        m = re.match(
                            r'^(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}-\d{2}-\d{2}|\d{1,2}\s+\w+\s+\d{2,4})\s+(.+?)\s+(\$?[\d,]+\.\d{2})$',
                            line
                        )
                        if m:
                            parsed_date = _parse_date(m.group(1))
                            merchant = m.group(2).strip()
                            amount = _parse_amount(m.group(3))
                            if parsed_date and merchant and amount and amount > 0:
                                is_dup = db.execute(
                                    select(models.Transaction).where(
                                        models.Transaction.date == datetime.date.fromisoformat(parsed_date),
                                        models.Transaction.merchant == merchant,
                                        func.round(models.Transaction.amount, 2) == round(amount, 2),
                                    )
                                ).first() is not None
                                suggested = _lookup_category(merchant, db)
                                rows_out.append({
                                    "date": parsed_date,
                                    "merchant": merchant,
                                    "amount": round(amount, 2),
                                    "suggested_category": suggested or "",
                                    "confidence": "high" if suggested else "low",
                                    "is_duplicate": is_dup,
                                })
    except Exception as e:
        raise HTTPException(400, f"Could not parse PDF: {str(e)}")

    # Deduplicate rows_out by (date, merchant, amount)
    seen = set()
    unique_rows = []
    for r in rows_out:
        key = (r["date"], r["merchant"], r["amount"])
        if key not in seen:
            seen.add(key)
            unique_rows.append(r)

    return {"rows": unique_rows, "pages": page_count}
