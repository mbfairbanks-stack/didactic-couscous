from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy.orm import Session
from sqlalchemy import select, func, distinct, text, nullslast
from sqlalchemy.exc import IntegrityError
from pydantic import BaseModel
from typing import Optional, List
import datetime
import hashlib
import tempfile, os, io, json, csv, re, math
from collections import defaultdict

from dotenv import load_dotenv
load_dotenv()

import models, database
from database import engine, get_db, run_migrations, DEMO_MODE, set_current_db_path
from importer import import_xlsx
import user_auth

models.Base.metadata.create_all(bind=engine)
run_migrations()

BUDGET_PASSWORD = os.getenv("BUDGET_PASSWORD", "")
MULTI_USER = os.getenv("MULTI_USER", "true").lower() not in ("0", "false", "no")

# Seed built-in users (demo + admin) on startup
if MULTI_USER:
    user_auth.seed_default_users(admin_password=BUDGET_PASSWORD)

app = FastAPI(title="BudgetBot API")

from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

_ALLOWED_ORIGINS = [o.strip() for o in os.getenv("ALLOWED_ORIGINS", "*").split(",")]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

_PUBLIC_PATHS = {"/auth/login", "/auth/register", "/auth/check"}


@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    if request.method == "OPTIONS":
        return await call_next(request)

    if MULTI_USER:
        # Allow login/register without a token
        if request.url.path in _PUBLIC_PATHS:
            return await call_next(request)

        token = request.headers.get("Authorization", "").removeprefix("Bearer ").strip()
        user = user_auth.get_user_from_token(token) if token else None
        if not user:
            return JSONResponse(status_code=401, content={"detail": "Unauthorized"})

        username, is_demo = user
        # Demo users cannot write data
        if is_demo and request.method in ("POST", "PUT", "PATCH", "DELETE"):
            return JSONResponse(
                status_code=403,
                content={"detail": "Demo mode is read-only. Create your own account to get started!"},
            )
        # Route this request to the correct user DB
        db_path = user_auth.get_user_db_path(username, is_demo)
        set_current_db_path(db_path)
        request.state.db_path = db_path
        request.state.username = username
        request.state.is_demo = is_demo
        return await call_next(request)

    # ── Legacy single-password mode (local dev / old deployments) ──
    if not BUDGET_PASSWORD:
        if DEMO_MODE and request.method in ("POST", "PUT", "PATCH", "DELETE"):
            if request.url.path not in _PUBLIC_PATHS:
                return JSONResponse(
                    status_code=403,
                    content={"detail": "Demo mode is read-only."},
                )
        return await call_next(request)

    if request.url.path in _PUBLIC_PATHS:
        return await call_next(request)

    pw_hash = hashlib.sha256(BUDGET_PASSWORD.encode()).hexdigest()
    token = request.headers.get("Authorization", "").removeprefix("Bearer ").strip()
    if token != pw_hash:
        return JSONResponse(status_code=401, content={"detail": "Unauthorized"})
    return await call_next(request)


# ---------------------------------------------------------------------------
# Auth endpoints
# ---------------------------------------------------------------------------

class LoginRequest(BaseModel):
    username: Optional[str] = None
    password: str


class RegisterRequest(BaseModel):
    username: str
    password: str


@app.post("/auth/register")
@limiter.limit("5/minute")
async def auth_register(request: Request, body: RegisterRequest):
    if not MULTI_USER:
        raise HTTPException(status_code=404)
    if len(body.username) < 3:
        raise HTTPException(status_code=400, detail="Username must be at least 3 characters")
    if len(body.password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    try:
        token = user_auth.register(body.username, body.password)
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))
    return {"token": token, "demo": False}


@app.post("/auth/login")
@limiter.limit("10/minute")
async def auth_login(request: Request, body: LoginRequest):
    if MULTI_USER:
        username = (body.username or "").strip()
        if not username:
            raise HTTPException(status_code=400, detail="Username is required")
        try:
            token, is_demo = user_auth.authenticate(username, body.password)
        except ValueError:
            raise HTTPException(status_code=401, detail="Incorrect username or password")
        return {"token": token, "demo": is_demo}

    # Legacy single-password mode
    if not BUDGET_PASSWORD:
        return {"token": ""}
    pw_hash = hashlib.sha256(BUDGET_PASSWORD.encode()).hexdigest()
    if hashlib.sha256(body.password.encode()).hexdigest() != pw_hash:
        raise HTTPException(status_code=401, detail="Incorrect password")
    return {"token": pw_hash, "demo": DEMO_MODE}


@app.get("/auth/check")
def auth_check(request: Request):
    """Returns ok:true only if the token is valid."""
    if MULTI_USER:
        token = request.headers.get("Authorization", "").removeprefix("Bearer ").strip()
        user = user_auth.get_user_from_token(token) if token else None
        if not user:
            return {"ok": False, "demo": False}
        _, is_demo = user
        return {"ok": True, "demo": is_demo}
    is_demo = getattr(request.state, "is_demo", DEMO_MODE)
    return {"ok": True, "demo": is_demo}


# ---------------------------------------------------------------------------
# Pydantic schemas
# ---------------------------------------------------------------------------

class TransactionCreate(BaseModel):
    date: datetime.date
    merchant: str
    amount: float
    category: str
    year: int
    month: int
    is_fixed: Optional[bool] = False
    is_recurring: Optional[bool] = False
    notes: Optional[str] = None
    source: Optional[str] = None
    linked_debt_id: Optional[int] = None
    debt_direction: Optional[str] = None  # "payment" or "charge"


class TransactionUpdate(BaseModel):
    date: Optional[datetime.date] = None
    merchant: Optional[str] = None
    amount: Optional[float] = None
    category: Optional[str] = None
    is_fixed: Optional[bool] = None
    is_recurring: Optional[bool] = None
    notes: Optional[str] = None
    linked_debt_id: Optional[int] = None
    debt_direction: Optional[str] = None


class TransactionOut(TransactionCreate):
    id: int
    model_config = {"from_attributes": True}


class IncomeCreate(BaseModel):
    year: int
    month: int
    person: str
    income_type: str
    amount: float
    pay_date: Optional[datetime.date] = None
    rrsp_employee: float = 0.0
    rrsp_employer: float = 0.0
    espp_deduction: float = 0.0


class IncomeOut(IncomeCreate):
    id: int
    model_config = {"from_attributes": True}


class BudgetTargetCreate(BaseModel):
    category: str
    year: int
    month: Optional[int] = None
    amount: float


class BudgetTargetOut(BudgetTargetCreate):
    id: int
    model_config = {"from_attributes": True}


# ---------------------------------------------------------------------------
# Transactions
# ---------------------------------------------------------------------------

@app.get("/transactions", response_model=list[TransactionOut])
def list_transactions(
    year: Optional[int] = None,
    month: Optional[int] = None,
    category: Optional[str] = None,
    source: Optional[str] = None,
    skip: int = 0,
    limit: int = Query(default=500, le=10000),
    db: Session = Depends(get_db),
):
    q = select(models.Transaction)
    if year:
        q = q.where(models.Transaction.year == year)
    if month:
        q = q.where(models.Transaction.month == month)
    if category:
        q = q.where(models.Transaction.category == category)
    if source:
        q = q.where(models.Transaction.source == source)
    q = q.order_by(models.Transaction.date.desc()).offset(skip).limit(limit)
    return db.execute(q).scalars().all()


@app.get("/transactions/export")
def export_transactions_csv(
    year: Optional[int] = None,
    month: Optional[int] = None,
    category: Optional[str] = None,
    source: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Export transactions as a CSV file."""
    q = select(models.Transaction)
    if year:
        q = q.where(models.Transaction.year == year)
    if month:
        q = q.where(models.Transaction.month == month)
    if category:
        q = q.where(models.Transaction.category == category)
    if source:
        q = q.where(models.Transaction.source == source)
    q = q.order_by(models.Transaction.date.desc())
    txns = db.execute(q).scalars().all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["date", "merchant", "amount", "category", "is_fixed", "is_recurring", "notes", "source"])
    for t in txns:
        writer.writerow([t.date, t.merchant, t.amount, t.category, t.is_fixed, t.is_recurring, t.notes, t.source])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=transactions.csv"},
    )


@app.post("/transactions/auto-categorize")
def auto_categorize_transactions(db: Session = Depends(get_db)):
    """Auto-categorize transactions where category is null or 'Uncategorized'."""
    # Get all uncategorized transactions
    uncategorized = db.execute(
        select(models.Transaction).where(
            (models.Transaction.category == None) | (models.Transaction.category == "Uncategorized")
        )
    ).scalars().all()

    if not uncategorized:
        return {"updated": 0, "skipped": 0}

    # Get unique merchants that need categorization
    merchants = list({txn.merchant for txn in uncategorized})

    # Single query: for each merchant, find the most common category
    from sqlalchemy import case
    rows = db.execute(
        select(
            models.Transaction.merchant,
            models.Transaction.category,
            func.count(models.Transaction.id).label("cnt"),
        )
        .where(
            models.Transaction.merchant.in_(merchants),
            models.Transaction.category != None,
            models.Transaction.category != "Uncategorized",
        )
        .group_by(models.Transaction.merchant, models.Transaction.category)
        .order_by(func.count(models.Transaction.id).desc())
    ).all()

    # Build merchant -> best category map (first row for each merchant wins due to ORDER BY)
    merchant_cat: dict = {}
    for row in rows:
        if row.merchant not in merchant_cat:
            merchant_cat[row.merchant] = row.category

    updated = 0
    skipped = 0
    for txn in uncategorized:
        cat = merchant_cat.get(txn.merchant)
        if cat:
            txn.category = cat
            updated += 1
        else:
            skipped += 1

    db.commit()
    return {"updated": updated, "skipped": skipped}


def _sync_year_month(data: dict) -> dict:
    """If a date is present, always derive year/month from it."""
    if "date" in data and data["date"]:
        try:
            d = data["date"]
            if hasattr(d, "year"):
                data["year"] = d.year
                data["month"] = d.month
            else:
                parsed = datetime.date.fromisoformat(str(d))
                data["year"] = parsed.year
                data["month"] = parsed.month
        except (ValueError, AttributeError):
            pass
    return data


@app.post("/transactions", response_model=TransactionOut, status_code=201)
def create_transaction(body: TransactionCreate, db: Session = Depends(get_db)):
    data = _sync_year_month(body.model_dump())
    txn = models.Transaction(**data)
    db.add(txn)
    db.commit()
    db.refresh(txn)
    return txn


@app.put("/transactions/{txn_id}", response_model=TransactionOut)
def update_transaction(txn_id: int, body: TransactionUpdate, db: Session = Depends(get_db)):
    txn = db.get(models.Transaction, txn_id)
    if not txn:
        raise HTTPException(404, "Transaction not found")
    data = _sync_year_month(body.model_dump(exclude_none=True))
    for field, val in data.items():
        setattr(txn, field, val)
    db.commit()
    db.refresh(txn)
    return txn


@app.delete("/transactions/{txn_id}", status_code=204)
def delete_transaction(txn_id: int, db: Session = Depends(get_db)):
    txn = db.get(models.Transaction, txn_id)
    if not txn:
        raise HTTPException(404, "Transaction not found")
    db.delete(txn)
    db.commit()


# ---------------------------------------------------------------------------
# Income
# ---------------------------------------------------------------------------

@app.get("/income", response_model=list[IncomeOut])
def list_income(
    year: Optional[int] = None,
    month: Optional[int] = None,
    db: Session = Depends(get_db),
):
    q = select(models.Income)
    if year:
        q = q.where(models.Income.year == year)
    if month:
        q = q.where(models.Income.month == month)
    q = q.order_by(models.Income.year, models.Income.month)
    return db.execute(q).scalars().all()


@app.post("/income", response_model=IncomeOut, status_code=201)
def create_income(body: IncomeCreate, db: Session = Depends(get_db)):
    existing = db.execute(
        select(models.Income).where(
            models.Income.year == body.year,
            models.Income.month == body.month,
            models.Income.person == body.person,
            models.Income.income_type == body.income_type,
            models.Income.pay_date == body.pay_date,
        )
    ).scalar_one_or_none()
    if existing:
        existing.amount = body.amount
        db.commit()
        db.refresh(existing)
        return existing
    income = models.Income(**body.model_dump())
    try:
        db.add(income)
        db.commit()
        db.refresh(income)
        return income
    except IntegrityError:
        db.rollback()
        raise HTTPException(409, "Duplicate entry — record already exists")


@app.put("/income/{income_id}", response_model=IncomeOut)
def update_income(income_id: int, body: IncomeCreate, db: Session = Depends(get_db)):
    income = db.get(models.Income, income_id)
    if not income:
        raise HTTPException(404, "Income record not found")
    for field, val in body.model_dump().items():
        setattr(income, field, val)
    db.commit()
    db.refresh(income)
    return income


@app.delete("/income/{income_id}", status_code=204)
def delete_income(income_id: int, db: Session = Depends(get_db)):
    income = db.get(models.Income, income_id)
    if not income:
        raise HTTPException(404, "Income record not found")
    db.delete(income)
    db.commit()


# ---------------------------------------------------------------------------
# Budget targets
# ---------------------------------------------------------------------------

@app.get("/budget-targets", response_model=list[BudgetTargetOut])
def list_budget_targets(
    year: Optional[int] = None,
    month: Optional[int] = None,
    db: Session = Depends(get_db),
):
    q = select(models.BudgetTarget)
    if year:
        q = q.where(models.BudgetTarget.year == year)
    if month is not None:
        q = q.where(models.BudgetTarget.month == month)
    return db.execute(q).scalars().all()


@app.post("/budget-targets", response_model=BudgetTargetOut, status_code=201)
def create_budget_target(body: BudgetTargetCreate, db: Session = Depends(get_db)):
    target = models.BudgetTarget(**body.model_dump())
    try:
        db.add(target)
        db.commit()
        db.refresh(target)
        return target
    except IntegrityError:
        db.rollback()
        raise HTTPException(409, "Duplicate entry — record already exists")


@app.put("/budget-targets/{target_id}", response_model=BudgetTargetOut)
def update_budget_target(target_id: int, body: BudgetTargetCreate, db: Session = Depends(get_db)):
    target = db.get(models.BudgetTarget, target_id)
    if not target:
        raise HTTPException(404, "Budget target not found")
    for field, val in body.model_dump().items():
        setattr(target, field, val)
    db.commit()
    db.refresh(target)
    return target


@app.delete("/budget-targets/{target_id}", status_code=204)
def delete_budget_target(target_id: int, db: Session = Depends(get_db)):
    target = db.get(models.BudgetTarget, target_id)
    if not target:
        raise HTTPException(404, "Budget target not found")
    db.delete(target)
    db.commit()


# ---------------------------------------------------------------------------
# Summary / Analytics
# ---------------------------------------------------------------------------

@app.get("/summary/monthly")
def monthly_summary(year: int, db: Session = Depends(get_db)):
    """Returns monthly totals for income and expenses for a given year."""
    # Expenses by month
    expense_rows = db.execute(
        select(
            models.Transaction.month,
            func.sum(models.Transaction.amount).label("total"),
        )
        .where(models.Transaction.year == year)
        .group_by(models.Transaction.month)
        .order_by(models.Transaction.month)
    ).all()

    # Income by month
    income_rows = db.execute(
        select(
            models.Income.month,
            func.sum(models.Income.amount).label("total"),
        )
        .where(models.Income.year == year)
        .group_by(models.Income.month)
        .order_by(models.Income.month)
    ).all()

    months = list(range(1, 13))
    expenses = {r.month: r.total for r in expense_rows}
    income = {r.month: r.total for r in income_rows}

    return [
        {
            "month": m,
            "income": income.get(m, 0),
            "expenses": expenses.get(m, 0),
            "balance": income.get(m, 0) - expenses.get(m, 0),
        }
        for m in months
    ]


@app.get("/summary/categories")
def category_summary(year: int, month: Optional[int] = None, db: Session = Depends(get_db)):
    """Returns spending totals by category."""
    q = (
        select(
            models.Transaction.category,
            func.sum(models.Transaction.amount).label("total"),
            func.count(models.Transaction.id).label("count"),
        )
        .where(models.Transaction.year == year)
        .group_by(models.Transaction.category)
        .order_by(func.sum(models.Transaction.amount).desc())
    )
    if month:
        q = q.where(models.Transaction.month == month)
    rows = db.execute(q).all()
    return [{"category": r.category, "total": round(r.total, 2), "count": r.count} for r in rows]


@app.get("/summary/totals")
def totals_summary(year: int, month: Optional[int] = None, through_month: Optional[int] = None, db: Session = Depends(get_db)):
    """Returns high-level totals: total income, total expenses, balance."""
    q_exp = select(func.sum(models.Transaction.amount)).where(models.Transaction.year == year)
    q_inc = select(func.sum(models.Income.amount)).where(models.Income.year == year)
    if month:
        q_exp = q_exp.where(models.Transaction.month == month)
        q_inc = q_inc.where(models.Income.month == month)
    elif through_month:
        q_exp = q_exp.where(models.Transaction.month <= through_month)
        q_inc = q_inc.where(models.Income.month <= through_month)
    total_expenses = db.execute(q_exp).scalar() or 0
    total_income = db.execute(q_inc).scalar() or 0
    return {
        "total_income": total_income,
        "total_expenses": total_expenses,
        "balance": total_income - total_expenses,
        "savings_rate": round((total_income - total_expenses) / total_income * 100, 1) if total_income else 0,
    }


@app.get("/summary/daily")
def daily_summary(year: int, month: Optional[int] = None, db: Session = Depends(get_db)):
    """Returns spending totals grouped by date for a given year/month."""
    q = (
        select(models.Transaction.date, func.sum(models.Transaction.amount).label("total"))
        .where(models.Transaction.year == year)
        .where(models.Transaction.amount > 0)
    )
    if month:
        q = q.where(models.Transaction.month == month)
    q = q.group_by(models.Transaction.date).order_by(models.Transaction.date)
    rows = db.execute(q).all()
    return [{"date": r.date, "total": round(r.total, 2)} for r in rows]


@app.get("/summary/category-trend")
def category_trend(category: str, db: Session = Depends(get_db)):
    """Returns monthly spending for a category across all years."""
    rows = db.execute(
        select(
            models.Transaction.year,
            models.Transaction.month,
            func.sum(models.Transaction.amount).label("total"),
        )
        .where(models.Transaction.category == category)
        .group_by(models.Transaction.year, models.Transaction.month)
        .order_by(models.Transaction.year, models.Transaction.month)
    ).all()
    return [{"year": r.year, "month": r.month, "total": round(r.total, 2)} for r in rows]


@app.get("/summary/projections")
def projections_summary(year: int, month: int, db: Session = Depends(get_db)):
    """Historical averages + year-end projections based on current run rate."""
    # Historical monthly income (all months except current)
    income_months = db.execute(
        select(models.Income.year, models.Income.month, func.sum(models.Income.amount).label("total"))
        .where(~((models.Income.year == year) & (models.Income.month == month)))
        .group_by(models.Income.year, models.Income.month)
    ).all()
    avg_monthly_income = (sum(r.total for r in income_months) / len(income_months)) if income_months else 0

    # Historical monthly expenses (all months except current)
    expense_months = db.execute(
        select(models.Transaction.year, models.Transaction.month, func.sum(models.Transaction.amount).label("total"))
        .where(~((models.Transaction.year == year) & (models.Transaction.month == month)))
        .group_by(models.Transaction.year, models.Transaction.month)
    ).all()
    avg_monthly_expenses = (sum(r.total for r in expense_months) / len(expense_months)) if expense_months else 0

    # Fixed expense categories from summary source — avg per category per month
    fixed_rows = db.execute(
        select(
            models.Transaction.category,
            models.Transaction.year,
            models.Transaction.month,
            func.sum(models.Transaction.amount).label("total"),
        )
        .where(models.Transaction.source == "summary")
        .group_by(models.Transaction.category, models.Transaction.year, models.Transaction.month)
    ).all()

    cat_monthly = defaultdict(list)
    for r in fixed_rows:
        cat_monthly[r.category].append(r.total)

    fixed_categories = sorted(
        [{"category": cat, "avg_monthly": round(sum(v) / len(v), 2)} for cat, v in cat_monthly.items()],
        key=lambda x: -x["avg_monthly"],
    )
    fixed_monthly_total = sum(c["avg_monthly"] for c in fixed_categories)

    # YTD actuals
    ytd_income = db.execute(
        select(func.sum(models.Income.amount))
        .where(models.Income.year == year, models.Income.month <= month)
    ).scalar() or 0

    ytd_expenses = db.execute(
        select(func.sum(models.Transaction.amount))
        .where(models.Transaction.year == year, models.Transaction.month <= month)
    ).scalar() or 0

    remaining = 12 - month
    projected_income = ytd_income + remaining * avg_monthly_income
    projected_expenses = ytd_expenses + remaining * avg_monthly_expenses

    return {
        "avg_monthly_income": round(avg_monthly_income, 2),
        "avg_monthly_expenses": round(avg_monthly_expenses, 2),
        "fixed_monthly_total": round(fixed_monthly_total, 2),
        "fixed_categories": fixed_categories,
        "projected_year_income": round(projected_income, 2),
        "projected_year_expenses": round(projected_expenses, 2),
        "projected_year_balance": round(projected_income - projected_expenses, 2),
    }


@app.get("/summary/forecast")
def forecast_summary(year: int, month: int, db: Session = Depends(get_db)):
    """Extrapolate month-end spending per category based on days elapsed."""
    from datetime import datetime
    import calendar

    days_in_month = calendar.monthrange(year, month)[1]
    now = datetime.now()
    if now.year == year and now.month == month:
        days_elapsed = min(now.day, days_in_month)
    else:
        days_elapsed = days_in_month

    rows = db.execute(
        select(
            models.Transaction.category,
            func.sum(models.Transaction.amount).label("actual"),
        )
        .where(models.Transaction.year == year, models.Transaction.month == month)
        .group_by(models.Transaction.category)
    ).all()

    result = []
    for r in rows:
        if r.actual and r.actual > 0:
            forecast = (r.actual / days_elapsed * days_in_month) if days_elapsed > 0 else r.actual
            result.append({
                "category": r.category,
                "actual": round(r.actual, 2),
                "forecast": round(forecast, 2),
                "days_elapsed": days_elapsed,
                "days_in_month": days_in_month,
            })

    return result


@app.get("/settings")
def get_settings(db: Session = Depends(get_db)):
    rows = db.execute(select(models.AppSettings)).scalars().all()
    return {r.key: r.value for r in rows}


@app.put("/settings")
def update_settings(body: dict, db: Session = Depends(get_db)):
    for key, value in body.items():
        existing = db.get(models.AppSettings, key)
        if existing:
            existing.value = str(value)
        else:
            db.add(models.AppSettings(key=key, value=str(value)))
    db.commit()
    return {"ok": True}


@app.get("/category-definitions")
def list_category_definitions(db: Session = Depends(get_db)):
    rows = db.execute(select(models.Category).order_by(models.Category.group_name, models.Category.name)).scalars().all()
    return [{"id": r.id, "name": r.name, "group": r.group_name, "is_legacy": bool(r.is_legacy),
             "is_hidden": bool(r.is_hidden), "parent_name": r.parent_name} for r in rows]


class CategoryCreate(BaseModel):
    name: str
    group: str
    parent_name: Optional[str] = None


class CategoryUpdate(BaseModel):
    name: Optional[str] = None
    group: Optional[str] = None
    is_hidden: Optional[bool] = None
    parent_name: Optional[str] = None


@app.post("/category-definitions", status_code=201)
def create_category_definition(body: CategoryCreate, db: Session = Depends(get_db)):
    existing = db.execute(select(models.Category).where(models.Category.name == body.name)).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=409, detail=f"Category '{body.name}' already exists")
    cat = models.Category(name=body.name, group_name=body.group, is_legacy=False, parent_name=body.parent_name)
    db.add(cat)
    db.commit()
    db.refresh(cat)
    return {"id": cat.id, "name": cat.name, "group": cat.group_name, "is_legacy": False,
            "is_hidden": False, "parent_name": cat.parent_name}


@app.put("/category-definitions/{cat_id}")
def update_category_definition(cat_id: int, body: CategoryUpdate, db: Session = Depends(get_db)):
    cat = db.get(models.Category, cat_id)
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    if body.name is not None:
        cat.name = body.name
    if body.group is not None:
        cat.group_name = body.group
    if body.is_hidden is not None:
        cat.is_hidden = body.is_hidden
    if body.parent_name is not None:
        cat.parent_name = body.parent_name
    elif "parent_name" in body.model_fields_set:
        cat.parent_name = None  # explicitly clearing it
    db.commit()
    return {"id": cat.id, "name": cat.name, "group": cat.group_name, "is_legacy": bool(cat.is_legacy),
            "is_hidden": bool(cat.is_hidden), "parent_name": cat.parent_name}


@app.delete("/category-definitions/{cat_id}", status_code=204)
def delete_category_definition(cat_id: int, db: Session = Depends(get_db)):
    cat = db.get(models.Category, cat_id)
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    count = db.execute(
        select(func.count()).where(models.Transaction.category == cat.name)
    ).scalar()
    if count > 0:
        raise HTTPException(status_code=409, detail=f"Cannot delete: {count} transaction(s) use this category")
    db.delete(cat)
    db.commit()


class CategoryMerge(BaseModel):
    source_names: list[str]
    target_name: str


@app.post("/category-definitions/merge")
def merge_categories(body: CategoryMerge, db: Session = Depends(get_db)):
    """Reassign all transactions from source categories to target, then hide sources."""
    # Validate target category exists
    target_exists = db.execute(
        select(models.Category).where(models.Category.name == body.target_name)
    ).scalar_one_or_none()
    if not target_exists:
        raise HTTPException(404, f"Target category '{body.target_name}' does not exist")
    updated = 0
    for name in body.source_names:
        result = db.execute(
            text("UPDATE transactions SET category = :target WHERE category = :source"),
            {"target": body.target_name, "source": name},
        )
        updated += result.rowcount
        # Hide the source category
        src = db.execute(select(models.Category).where(models.Category.name == name)).scalar_one_or_none()
        if src:
            src.is_hidden = True
    db.commit()
    return {"merged_transactions": updated, "hidden": body.source_names}


@app.get("/onboarding-status")
def onboarding_status(db: Session = Depends(get_db)):
    setting = db.get(models.AppSettings, "onboarding_complete")
    has_transactions = db.execute(select(func.count()).select_from(models.Transaction)).scalar() > 0
    p1 = db.get(models.AppSettings, "person_1")
    p2 = db.get(models.AppSettings, "person_2")
    has_people = (p1 and p1.value not in ("Person 1", "")) or (p2 and p2.value not in ("Person 2", ""))
    needs_onboarding = not (setting and setting.value == "true")
    return {
        "needs_onboarding": needs_onboarding,
        "has_transactions": has_transactions,
        "has_people_configured": has_people,
    }


@app.get("/categories")
def list_categories(db: Session = Depends(get_db)):
    rows = db.execute(
        select(distinct(models.Transaction.category))
        .where(models.Transaction.category != None)  # noqa: E711
        .where(models.Transaction.category != "")
        .order_by(models.Transaction.category)
    ).scalars().all()
    return rows


@app.get("/years")
def list_years(db: Session = Depends(get_db)):
    rows = db.execute(
        select(distinct(models.Transaction.year)).order_by(models.Transaction.year)
    ).scalars().all()
    return rows


# ---------------------------------------------------------------------------
# Deduplicate
# ---------------------------------------------------------------------------

BANK_ONLY_CATEGORIES = {
    "Mortgage", "Gas (Utility)", "Hydro",
    "Municipal Taxes", "Debt Payment", "Insurance", "Reliance",
}

@app.post("/cleanup-summary")
def cleanup_summary(db: Session = Depends(get_db)):
    """Remove all summary-source rows that aren't bank-only expenses.

    This removes any CC-covered categories that were incorrectly double-imported
    from the Summary tab (e.g. Groceries, Mobile, Internet, Security).
    """
    txns = db.execute(
        select(models.Transaction).where(models.Transaction.source == "summary")
    ).scalars().all()
    removed = 0
    for txn in txns:
        merchant_lower = (txn.merchant or "").strip().lower()
        is_bad_label = (
            "income" in merchant_lower or "saving" in merchant_lower
            or merchant_lower.startswith("total") or merchant_lower == "balance"
        )
        is_cc_covered = txn.category not in BANK_ONLY_CATEGORIES
        if is_bad_label or is_cc_covered:
            db.delete(txn)
            removed += 1
    db.commit()
    return {"removed": removed}


@app.post("/deduplicate")
def deduplicate_transactions(db: Session = Depends(get_db)):
    """
    Remove duplicate transactions keeping the lowest id for each
    (date, merchant, amount, category, year, month) group.
    """
    all_txns = db.execute(
        select(models.Transaction).order_by(models.Transaction.id)
    ).scalars().all()

    seen = {}
    to_delete = []
    for txn in all_txns:
        key = (txn.date, txn.merchant.strip().lower(), round(txn.amount, 2), txn.category, txn.year, txn.month)
        if key in seen:
            to_delete.append(txn)
        else:
            seen[key] = txn.id

    for txn in to_delete:
        db.delete(txn)
    db.commit()

    return {"duplicates_removed": len(to_delete)}


@app.get("/categories/suggested-renames")
def get_suggested_renames(db: Session = Depends(get_db)):
    """Return suggested renames filtered to only old categories that still exist in transactions."""
    from categories import SUGGESTED_RENAMES
    result = db.execute(text("SELECT DISTINCT category FROM transactions"))
    existing = {row[0] for row in result.fetchall()}
    return [
        {"from_category": k, "to_category": v}
        for k, v in SUGGESTED_RENAMES.items()
        if k in existing
    ]


@app.post("/categories/migrate")
def migrate_categories(db: Session = Depends(get_db)):
    """Apply all suggested category renames to existing transactions."""
    from categories import SUGGESTED_RENAMES
    total = 0
    for old, new in SUGGESTED_RENAMES.items():
        result = db.execute(
            text("UPDATE transactions SET category = :new WHERE category = :old"),
            {"new": new, "old": old}
        )
        total += result.rowcount
    db.commit()
    return {"updated": total}


# ---------------------------------------------------------------------------
# Debts
# ---------------------------------------------------------------------------

def _compute_debt_balance(
    initial_balance: float,
    annual_rate: float,
    initial_date_str: Optional[str],
    linked_txns: list,
) -> Optional[float]:
    """Walk month-by-month from initial_date, applying interest and linked transactions."""
    if not initial_date_str:
        return None
    try:
        start = datetime.date.fromisoformat(initial_date_str)
    except ValueError:
        return None

    monthly_rate = annual_rate / 12.0
    balance = float(initial_balance)
    today = datetime.date.today()

    # Sort transactions by date
    sorted_txns = sorted(linked_txns, key=lambda t: t.date)
    tx_idx = 0

    # Walk month by month from the start month through the current month
    year, month = start.year, start.month
    while (year, month) <= (today.year, today.month):
        # End of this calendar month
        if month == 12:
            next_year, next_month = year + 1, 1
        else:
            next_year, next_month = year, month + 1

        month_start = datetime.date(year, month, 1)
        month_end = datetime.date(next_year, next_month, 1)

        # Apply monthly interest (skip first partial month if start_date is mid-month)
        if monthly_rate > 0:
            balance += balance * monthly_rate

        # Apply transactions in this calendar month
        while tx_idx < len(sorted_txns):
            t = sorted_txns[tx_idx]
            t_date = t.date if isinstance(t.date, datetime.date) else datetime.date.fromisoformat(str(t.date))
            if t_date >= month_end:
                break
            if t_date >= month_start and t_date >= start:
                if t.debt_direction == "charge":
                    balance += t.amount
                else:  # "payment" or None defaults to payment
                    balance -= t.amount
            tx_idx += 1

        year, month = next_year, next_month

    return round(max(balance, 0.0), 2)


class DebtCreate(BaseModel):
    name: str
    creditor: str
    debt_type: str = "loan"         # "loan", "loc", or "mortgage"
    credit_limit: float = 0.0       # LOC only
    interest_rate: float = 0.0      # annual rate, e.g. 0.0645
    initial_balance: float = 0.0
    current_balance: float = 0.0
    monthly_payment: float = 0.0
    monthly_extra: float = 0.0
    savings: float = 0.0
    due_date: Optional[str] = None
    notes: Optional[str] = None
    linked_asset_id: Optional[int] = None
    initial_date: Optional[str] = None


class DebtOut(BaseModel):
    id: int
    name: str
    creditor: str
    debt_type: str
    credit_limit: float
    interest_rate: float
    initial_balance: float
    current_balance: float
    monthly_payment: float
    monthly_extra: float
    savings: float
    due_date: Optional[str]
    notes: Optional[str]
    linked_asset_id: Optional[int]
    initial_date: Optional[str] = None
    equity: Optional[float] = None  # computed: linked asset balance - current_balance
    computed_balance: Optional[float] = None
    model_config = {"from_attributes": True}


@app.get("/debts", response_model=list[DebtOut])
def list_debts(db: Session = Depends(get_db)):
    debts = db.execute(select(models.Debt).order_by(models.Debt.name)).scalars().all()
    result = []
    for debt in debts:
        d = DebtOut.model_validate(debt)
        if debt.debt_type == "mortgage" and debt.linked_asset_id:
            asset = db.get(models.Asset, debt.linked_asset_id)
            if asset:
                d.equity = asset.balance - debt.current_balance
        # Compute balance from linked transactions if initial_date is set
        if debt.initial_date:
            linked_txns = db.execute(
                select(models.Transaction).where(models.Transaction.linked_debt_id == debt.id)
            ).scalars().all()
            d.computed_balance = _compute_debt_balance(
                debt.initial_balance, debt.interest_rate, debt.initial_date, linked_txns
            )
            if d.computed_balance is not None and debt.debt_type == "mortgage" and debt.linked_asset_id:
                asset = db.get(models.Asset, debt.linked_asset_id)
                if asset:
                    d.equity = asset.balance - d.computed_balance
        result.append(d)
    return result


@app.post("/debts", response_model=DebtOut, status_code=201)
def create_debt(body: DebtCreate, db: Session = Depends(get_db)):
    debt = models.Debt(**body.model_dump())
    db.add(debt)
    db.commit()
    db.refresh(debt)
    return debt


@app.put("/debts/{debt_id}", response_model=DebtOut)
def update_debt(debt_id: int, body: DebtCreate, db: Session = Depends(get_db)):
    debt = db.get(models.Debt, debt_id)
    if not debt:
        raise HTTPException(404, "Debt not found")
    for field, val in body.model_dump().items():
        setattr(debt, field, val)
    db.commit()
    db.refresh(debt)
    return debt


@app.delete("/debts/{debt_id}", status_code=204)
def delete_debt(debt_id: int, db: Session = Depends(get_db)):
    debt = db.get(models.Debt, debt_id)
    if not debt:
        raise HTTPException(404, "Debt not found")
    db.delete(debt)
    db.commit()


@app.get("/debts/{debt_id}/transactions")
def get_debt_transactions(debt_id: int, db: Session = Depends(get_db)):
    """Return all transactions linked to a debt, with running balance."""
    debt = db.get(models.Debt, debt_id)
    if not debt:
        raise HTTPException(404, "Debt not found")

    txns = db.execute(
        select(models.Transaction)
        .where(models.Transaction.linked_debt_id == debt_id)
        .order_by(models.Transaction.date)
    ).scalars().all()

    return [
        {
            "id": t.id,
            "date": t.date.isoformat() if hasattr(t.date, "isoformat") else str(t.date),
            "merchant": t.merchant,
            "amount": t.amount,
            "debt_direction": t.debt_direction or "payment",
            "category": t.category,
            "notes": t.notes,
        }
        for t in txns
    ]


# ---------------------------------------------------------------------------
# Assets (DB-backed net worth assets)
# ---------------------------------------------------------------------------

class AssetCreate(BaseModel):
    name: str
    asset_type: str = "other"
    balance: float = 0.0
    notes: Optional[str] = None
    sort_order: int = 0
    auto_sync: bool = False
    liquidity: str = "liquid"  # "liquid" or "illiquid"


class AssetOut(AssetCreate):
    id: int
    model_config = {"from_attributes": True}


ASSET_DEFAULTS = [
    {"name": "Checking", "asset_type": "cash", "sort_order": 1, "auto_sync": False, "liquidity": "liquid"},
    {"name": "Savings", "asset_type": "cash", "sort_order": 2, "auto_sync": False, "liquidity": "liquid"},
    {"name": "RRSP (Payroll)", "asset_type": "rrsp", "sort_order": 3, "auto_sync": True, "liquidity": "illiquid"},
    {"name": "TFSA", "asset_type": "tfsa", "sort_order": 4, "auto_sync": False, "liquidity": "liquid"},
    {"name": "ESPP (Block)", "asset_type": "espp", "sort_order": 5, "auto_sync": True, "liquidity": "liquid"},
]


@app.get("/assets", response_model=list[AssetOut])
def list_assets(db: Session = Depends(get_db)):
    rows = db.execute(select(models.Asset).order_by(models.Asset.sort_order, models.Asset.name)).scalars().all()
    if not rows:
        # Seed defaults on first call
        for i, d in enumerate(ASSET_DEFAULTS):
            db.add(models.Asset(**d, balance=0.0))
        db.commit()
        rows = db.execute(select(models.Asset).order_by(models.Asset.sort_order, models.Asset.name)).scalars().all()
    return rows


@app.post("/assets", response_model=AssetOut, status_code=201)
def create_asset(body: AssetCreate, db: Session = Depends(get_db)):
    asset = models.Asset(**body.model_dump())
    db.add(asset)
    db.commit()
    db.refresh(asset)
    return asset


@app.put("/assets/{asset_id}", response_model=AssetOut)
def update_asset(asset_id: int, body: AssetCreate, db: Session = Depends(get_db)):
    asset = db.get(models.Asset, asset_id)
    if not asset:
        raise HTTPException(404, "Asset not found")
    for field, val in body.model_dump().items():
        setattr(asset, field, val)
    db.commit()
    db.refresh(asset)
    return asset


@app.delete("/assets/{asset_id}", status_code=204)
def delete_asset(asset_id: int, db: Session = Depends(get_db)):
    asset = db.get(models.Asset, asset_id)
    if not asset:
        raise HTTPException(404, "Asset not found")
    db.delete(asset)
    db.commit()


@app.post("/assets/sync-savings")
def sync_savings_assets(db: Session = Depends(get_db)):
    """Auto-update RRSP and ESPP asset balances from tracked contribution data."""
    # All-time RRSP total (employee + employer contributions from income records)
    rrsp_total = db.execute(
        select(
            func.sum(models.Income.rrsp_employee + models.Income.rrsp_employer)
        )
    ).scalar() or 0

    # All-time ESPP deductions from income records (payroll deductions, not stock purchases)
    espp_total = db.execute(
        select(func.sum(models.Income.espp_deduction))
    ).scalar() or 0

    updated = []
    assets = db.execute(select(models.Asset).where(models.Asset.auto_sync == True)).scalars().all()
    for asset in assets:
        if asset.asset_type == "rrsp" and rrsp_total > 0:
            asset.balance = round(rrsp_total, 2)
            updated.append({"id": asset.id, "name": asset.name, "balance": asset.balance})
        elif asset.asset_type == "espp" and espp_total > 0:
            asset.balance = round(espp_total, 2)
            updated.append({"id": asset.id, "name": asset.name, "balance": asset.balance})
    db.commit()
    return {"updated": updated, "rrsp_total": round(rrsp_total, 2), "espp_total": round(espp_total, 2)}


# ---------------------------------------------------------------------------
# Savings contributions (RRSP + ESPP per paycheck)
# ---------------------------------------------------------------------------

RRSP_ANNUAL_MAX = 8400.0  # employee portion; with 50% employer match total = $12,600
RRSP_MATCH_RATE = 0.50
ESPP_DEDUCTION_RATE = 0.10
ESPP_DISCOUNT_RATE = 0.15


class SavingsContributionCreate(BaseModel):
    pay_date: str
    year: int
    month: int
    gross_income: float
    rrsp_employee: float = 0.0
    rrsp_employer: Optional[float] = None  # auto-calculated if None
    espp_deduction: Optional[float] = None  # auto-calculated if None
    notes: Optional[str] = None


class SavingsContributionOut(SavingsContributionCreate):
    id: int
    model_config = {"from_attributes": True}


class EsppPurchaseCreate(BaseModel):
    purchase_date: str
    period_start: Optional[str] = None
    period_end: Optional[str] = None
    total_deducted: float = 0.0
    shares_purchased: float = 0.0
    purchase_price: float = 0.0
    market_price: float = 0.0
    current_price: float = 0.0
    notes: Optional[str] = None


class EsppPurchaseOut(EsppPurchaseCreate):
    id: int
    model_config = {"from_attributes": True}


@app.get("/savings-contributions", response_model=list[SavingsContributionOut])
def list_savings_contributions(year: Optional[int] = None, db: Session = Depends(get_db)):
    q = select(models.SavingsContribution).order_by(models.SavingsContribution.pay_date.desc())
    if year:
        q = q.where(models.SavingsContribution.year == year)
    return db.execute(q).scalars().all()


@app.post("/savings-contributions", response_model=SavingsContributionOut, status_code=201)
def create_savings_contribution(body: SavingsContributionCreate, db: Session = Depends(get_db)):
    data = body.model_dump()
    # Auto-calculate employer match and ESPP if not provided
    if data["rrsp_employer"] is None:
        data["rrsp_employer"] = round(data["rrsp_employee"] * RRSP_MATCH_RATE, 2)
    if data["espp_deduction"] is None:
        data["espp_deduction"] = round(data["gross_income"] * ESPP_DEDUCTION_RATE, 2)
    contrib = models.SavingsContribution(**data)
    db.add(contrib)
    db.commit()
    db.refresh(contrib)
    return contrib


@app.put("/savings-contributions/{contrib_id}", response_model=SavingsContributionOut)
def update_savings_contribution(contrib_id: int, body: SavingsContributionCreate, db: Session = Depends(get_db)):
    contrib = db.get(models.SavingsContribution, contrib_id)
    if not contrib:
        raise HTTPException(404, "Contribution not found")
    data = body.model_dump()
    if data["rrsp_employer"] is None:
        data["rrsp_employer"] = round(data["rrsp_employee"] * RRSP_MATCH_RATE, 2)
    if data["espp_deduction"] is None:
        data["espp_deduction"] = round(data["gross_income"] * ESPP_DEDUCTION_RATE, 2)
    for field, val in data.items():
        setattr(contrib, field, val)
    db.commit()
    db.refresh(contrib)
    return contrib


@app.delete("/savings-contributions/{contrib_id}", status_code=204)
def delete_savings_contribution(contrib_id: int, db: Session = Depends(get_db)):
    contrib = db.get(models.SavingsContribution, contrib_id)
    if not contrib:
        raise HTTPException(404, "Contribution not found")
    db.delete(contrib)
    db.commit()


@app.get("/savings-contributions/summary")
def savings_summary(year: int, db: Session = Depends(get_db)):
    """YTD RRSP + ESPP totals sourced from income records, plus cap progress with carryover."""
    rows = db.execute(
        select(models.Income).where(models.Income.year == year)
    ).scalars().all()

    ytd_rrsp_employee = sum(r.rrsp_employee for r in rows)
    ytd_rrsp_employer = sum(r.rrsp_employer for r in rows)
    ytd_rrsp_total = ytd_rrsp_employee + ytd_rrsp_employer
    ytd_espp = sum(r.espp_deduction for r in rows)

    # Count distinct pay dates that have any savings activity
    pay_dates_with_savings = len({r.pay_date for r in rows if (r.rrsp_employee or 0) > 0 or (r.espp_deduction or 0) > 0})

    # RRSP carryover: unused contribution room from prior years accumulates
    prior_year_totals = db.execute(
        select(models.Income.year, func.sum(models.Income.rrsp_employee).label("total"))
        .where(models.Income.year < year)
        .group_by(models.Income.year)
    ).all()
    carryover_room = sum(max(RRSP_ANNUAL_MAX - (row.total or 0), 0) for row in prior_year_totals)
    effective_rrsp_cap = RRSP_ANNUAL_MAX + carryover_room

    # ESPP: all-time deductions vs total used in purchases (to show pending balance)
    all_espp_deducted = db.execute(
        select(func.sum(models.Income.espp_deduction))
    ).scalar() or 0
    all_espp_purchased = db.execute(
        select(func.sum(models.EsppPurchase.total_deducted))
    ).scalar() or 0
    espp_pending = max(all_espp_deducted - all_espp_purchased, 0)

    # ESPP purchases this year
    purchases = db.execute(
        select(models.EsppPurchase).where(
            models.EsppPurchase.purchase_date.like(f"{year}%")
        )
    ).scalars().all()
    espp_current_value = sum(
        p.shares_purchased * (p.current_price or p.market_price)
        for p in purchases
    )

    return {
        "year": year,
        "rrsp_employee_ytd": round(ytd_rrsp_employee, 2),
        "rrsp_employer_ytd": round(ytd_rrsp_employer, 2),
        "rrsp_total_ytd": round(ytd_rrsp_total, 2),
        "rrsp_annual_max": RRSP_ANNUAL_MAX,
        "rrsp_carryover": round(carryover_room, 2),
        "rrsp_effective_cap": round(effective_rrsp_cap, 2),
        "rrsp_remaining": round(max(effective_rrsp_cap - ytd_rrsp_employee, 0), 2),
        "rrsp_pct": round(min(ytd_rrsp_employee / effective_rrsp_cap * 100, 100) if effective_rrsp_cap > 0 else 0, 1),
        "espp_deducted_ytd": round(ytd_espp, 2),
        "espp_pending_all_time": round(espp_pending, 2),
        "espp_current_value": round(espp_current_value, 2),
        "espp_discount_rate": ESPP_DISCOUNT_RATE,
        "rrsp_match_rate": RRSP_MATCH_RATE,
        "contributions": pay_dates_with_savings,
    }


# ESPP Purchases
@app.get("/espp-purchases", response_model=list[EsppPurchaseOut])
def list_espp_purchases(db: Session = Depends(get_db)):
    return db.execute(select(models.EsppPurchase).order_by(models.EsppPurchase.purchase_date.desc())).scalars().all()


@app.post("/espp-purchases", response_model=EsppPurchaseOut, status_code=201)
def create_espp_purchase(body: EsppPurchaseCreate, db: Session = Depends(get_db)):
    purchase = models.EsppPurchase(**body.model_dump())
    db.add(purchase)
    db.commit()
    db.refresh(purchase)
    return purchase


@app.put("/espp-purchases/{purchase_id}", response_model=EsppPurchaseOut)
def update_espp_purchase(purchase_id: int, body: EsppPurchaseCreate, db: Session = Depends(get_db)):
    purchase = db.get(models.EsppPurchase, purchase_id)
    if not purchase:
        raise HTTPException(404, "Purchase not found")
    for field, val in body.model_dump().items():
        setattr(purchase, field, val)
    db.commit()
    db.refresh(purchase)
    return purchase


@app.delete("/espp-purchases/{purchase_id}", status_code=204)
def delete_espp_purchase(purchase_id: int, db: Session = Depends(get_db)):
    purchase = db.get(models.EsppPurchase, purchase_id)
    if not purchase:
        raise HTTPException(404, "Purchase not found")
    db.delete(purchase)
    db.commit()


# ---------------------------------------------------------------------------
# Category Audit
# ---------------------------------------------------------------------------

@app.get("/merchant-categories")
def merchant_categories(db: Session = Depends(get_db)):
    """Return merchants assigned to more than one category (inconsistencies)."""
    rows = db.execute(
        select(
            models.Transaction.merchant,
            models.Transaction.category,
            func.count(models.Transaction.id).label("cnt"),
        )
        .group_by(models.Transaction.merchant, models.Transaction.category)
        .order_by(models.Transaction.merchant)
    ).all()

    by_merchant: dict[str, dict] = defaultdict(lambda: {"categories": {}, "count": 0})
    for r in rows:
        by_merchant[r.merchant]["categories"][r.category] = r.cnt
        by_merchant[r.merchant]["count"] += r.cnt

    result = []
    for merchant, data in sorted(by_merchant.items()):
        if len(data["categories"]) > 1:
            most_common = max(data["categories"], key=data["categories"].get)
            result.append({
                "merchant": merchant,
                "categories": list(data["categories"].keys()),
                "count": data["count"],
                "most_common": most_common,
            })
    return result


class BulkCategoryUpdate(BaseModel):
    merchant: str
    from_category: str
    to_category: str


@app.post("/transactions/bulk-category")
def bulk_update_category(body: BulkCategoryUpdate, db: Session = Depends(get_db)):
    """Reassign all transactions for a merchant from one category to another."""
    txns = db.execute(
        select(models.Transaction).where(
            models.Transaction.merchant == body.merchant,
            models.Transaction.category == body.from_category,
        )
    ).scalars().all()
    for txn in txns:
        txn.category = body.to_category
    db.commit()
    return {"updated": len(txns)}


class SetMerchantCategoryRequest(BaseModel):
    merchant: str
    category: str

@app.post("/transactions/set-merchant-category")
def set_merchant_category(body: SetMerchantCategoryRequest, db: Session = Depends(get_db)):
    """Reassign ALL transactions for a merchant to a single category."""
    result = db.execute(
        text("UPDATE transactions SET category = :cat WHERE merchant = :merchant"),
        {"cat": body.category, "merchant": body.merchant}
    )
    db.commit()
    return {"updated": result.rowcount}


# ---------------------------------------------------------------------------
# Smart CSV paste import
# ---------------------------------------------------------------------------

class ParseCsvRequest(BaseModel):
    text: str
    format: str = "auto"  # "auto", "amex", "td", "rbc", "cibc"


def _parse_amount(s: str) -> Optional[float]:
    if not s:
        return None
    cleaned = re.sub(r'[$,\s]', '', str(s))
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_date(s: str) -> Optional[str]:
    from datetime import datetime
    for fmt in ('%m/%d/%Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%y', '%Y/%m/%d',
                '%b %d, %Y', '%B %d, %Y', '%b %d %Y', '%B %d %Y',
                '%d-%b-%Y', '%d %b %Y', '%Y%m%d',
                '%d %b %y', '%d %B %y'):  # e.g. "19 May 26"
        try:
            return datetime.strptime(s.strip(), fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def _lookup_category(merchant: str, db: Session) -> Optional[str]:
    """Find category: check merchant_rules first, then transaction history."""
    # 1. Exact match in merchant_rules
    rule = db.execute(
        select(models.MerchantRule).where(models.MerchantRule.merchant_pattern == merchant)
    ).scalar_one_or_none()
    if rule:
        return rule.category

    # 2. Partial match in merchant_rules
    rules = db.execute(select(models.MerchantRule)).scalars().all()
    merchant_lower = merchant.lower()
    for r in rules:
        if r.merchant_pattern.lower() in merchant_lower or merchant_lower in r.merchant_pattern.lower():
            return r.category

    # 3. Exact match in transaction history
    rows = db.execute(
        select(models.Transaction.category, func.count(models.Transaction.id).label("cnt"))
        .where(models.Transaction.merchant == merchant)
        .group_by(models.Transaction.category)
        .order_by(func.count(models.Transaction.id).desc())
    ).all()
    if rows:
        return rows[0].category

    # 4. Fuzzy match in transaction history
    words = [w for w in re.sub(r'[^a-z\s]', '', merchant.lower()).split() if len(w) > 3]
    if not words:
        return None
    search_word = words[0]
    rows = db.execute(
        select(models.Transaction.category, func.count(models.Transaction.id).label("cnt"))
        .where(func.lower(models.Transaction.merchant).contains(search_word))
        .group_by(models.Transaction.category)
        .order_by(func.count(models.Transaction.id).desc())
    ).all()
    return rows[0].category if rows else None


@app.post("/parse-csv")
def parse_csv(body: ParseCsvRequest, db: Session = Depends(get_db)):
    """Parse CSV from AMEX/Visa/MC and return rows with suggested categories.
    Handles both headered (comma) and headerless tab-delimited (TD) formats.
    """
    raw = body.text.strip()
    if not raw:
        return {"rows": []}

    lines = [l.strip() for l in raw.splitlines() if l.strip()]

    # Skip a header row if present (e.g. "Date  Description  Amount")
    if lines and re.search(r'\bdate\b', lines[0], re.IGNORECASE) and re.search(r'\bdescription\b', lines[0], re.IGNORECASE):
        lines = lines[1:]

    # ── Detect AMEX multi-line paste format ──────────────────────────────────
    # Pattern: groups of 3 lines — date, merchant, $amount (repeating)
    # e.g.:
    #   19 May 26
    #   COUNTRY PAWS DOGPLEX 00 LONDON
    #   $6.10
    def _is_amex_multiline(lines: list) -> bool:
        if len(lines) < 3:
            return False
        # Check first few groups — allow remainder lines to be off
        matches = 0
        for i in range(0, min(len(lines) - 2, 12), 3):
            if _parse_date(lines[i]) and re.match(r'^-?\$?[\d,]+\.?\d*$', lines[i + 2]):
                matches += 1
        return matches >= 1

    if _is_amex_multiline(lines):
        parsed = []
        for i in range(0, len(lines) - 2, 3):
            date_str = lines[i].strip()
            merchant = lines[i + 1].strip()
            amt_str = lines[i + 2].strip()
            parsed_date = _parse_date(date_str)
            amount = _parse_amount(amt_str)
            if not parsed_date or not merchant or not amount or amount <= 0:
                continue
            suggested = _lookup_category(merchant, db)
            # Check if already in transactions
            is_dup = db.execute(
                select(models.Transaction).where(
                    models.Transaction.date == datetime.date.fromisoformat(parsed_date),
                    models.Transaction.merchant == merchant,
                    func.round(models.Transaction.amount, 2) == round(amount, 2),
                )
            ).first() is not None
            parsed.append({
                "date": parsed_date,
                "merchant": merchant,
                "amount": round(amount, 2),
                "suggested_category": suggested or "",
                "confidence": "high" if suggested else "low",
                "is_duplicate": is_dup,
            })
        return {"rows": parsed}

    # Detect headerless format: tab or comma delimited, first cell is a date
    # TD format: date,merchant,debit,credit,balance  (or tab-separated)
    def _split_first(line: str):
        if '\t' in line:
            return [p.strip() for p in line.split('\t')]
        return next(csv.reader([line]))

    def _is_headerless(line: str) -> bool:
        parts = _split_first(line)
        return len(parts) >= 3 and bool(_parse_date(parts[0]))

    use_td = bool(lines) and _is_headerless(lines[0])

    parsed = []

    if use_td:
        for line in lines:
            if not line.strip():
                continue
            parts = _split_first(line)
            if len(parts) < 3:
                continue
            date_str = parts[0]
            merchant = parts[1]
            # col 2 = debit (expense), col 3 = credit (payment), col 4 = balance
            debit = _parse_amount(parts[2]) if len(parts) > 2 else None
            credit = _parse_amount(parts[3]) if len(parts) > 3 else None

            parsed_date = _parse_date(date_str)
            if not parsed_date or not merchant:
                continue

            # Only import debits (expenses); skip credits/payments
            if debit and debit > 0:
                amount = debit
            else:
                continue

            suggested = _lookup_category(merchant, db)
            # Check if already in transactions
            is_dup = db.execute(
                select(models.Transaction).where(
                    models.Transaction.date == datetime.date.fromisoformat(parsed_date),
                    models.Transaction.merchant == merchant,
                    func.round(models.Transaction.amount, 2) == round(amount, 2),
                )
            ).first() is not None
            parsed.append({
                "date": parsed_date,
                "merchant": merchant,
                "amount": round(amount, 2),
                "suggested_category": suggested or "",
                "confidence": "high" if suggested else "low",
                "is_duplicate": is_dup,
            })
    else:
        # Header-based CSV (AMEX, RBC, CIBC, etc.)
        reader = csv.reader(io.StringIO(raw))
        rows = list(reader)
        if len(rows) < 2:
            return {"rows": []}

        header = [h.strip().lower().replace(' ', '_') for h in rows[0]]

        def col(names):
            for n in names:
                for i, h in enumerate(header):
                    if n in h:
                        return i
            return None

        date_col  = col(['date', 'transaction_date'])
        desc_col  = col(['description', 'desc', 'merchant', 'name', 'payee'])
        amt_col   = col(['amount'])
        debit_col = col(['debit'])
        cad_col   = col(['cad$', 'cad'])

        for row in rows[1:]:
            if not row or all(not c.strip() for c in row):
                continue
            merchant = row[desc_col].strip() if desc_col is not None and desc_col < len(row) else ""
            date_str = row[date_col].strip() if date_col is not None and date_col < len(row) else ""
            if not merchant or not date_str:
                continue

            amount = None
            if amt_col is not None and amt_col < len(row):
                amount = _parse_amount(row[amt_col])
            if (amount is None or amount <= 0) and debit_col is not None and debit_col < len(row):
                amount = _parse_amount(row[debit_col])
            if (amount is None or amount <= 0) and cad_col is not None and cad_col < len(row):
                amount = _parse_amount(row[cad_col])

            if amount is None or amount <= 0:
                continue

            parsed_date = _parse_date(date_str)
            if not parsed_date:
                continue

            suggested = _lookup_category(merchant, db)
            # Check if already in transactions
            is_dup = db.execute(
                select(models.Transaction).where(
                    models.Transaction.date == datetime.date.fromisoformat(parsed_date),
                    models.Transaction.merchant == merchant,
                    func.round(models.Transaction.amount, 2) == round(amount, 2),
                )
            ).first() is not None
            parsed.append({
                "date": parsed_date,
                "merchant": merchant,
                "amount": round(amount, 2),
                "suggested_category": suggested or "",
                "confidence": "high" if suggested else "low",
                "is_duplicate": is_dup,
            })

    return {"rows": parsed}


class CsvImportRow(BaseModel):
    date: str
    merchant: str
    amount: float
    category: str
    source: Optional[str] = None


@app.post("/import-csv-rows")
def import_csv_rows(rows: List[CsvImportRow], db: Session = Depends(get_db)):
    """Import reviewed CSV rows into the transactions table, skipping exact duplicates."""
    from datetime import datetime
    imported = 0
    skipped_duplicates = 0
    for row in rows:
        try:
            d = datetime.strptime(row.date, '%Y-%m-%d').date()
        except ValueError:
            continue
        dup = db.execute(
            select(models.Transaction).where(
                models.Transaction.date == d,
                models.Transaction.merchant == row.merchant,
                func.round(models.Transaction.amount, 2) == round(row.amount, 2),
            )
        ).first()
        if dup:
            skipped_duplicates += 1
            continue
        source = getattr(row, 'source', None) or "csv_import"
        txn = models.Transaction(
            date=d,
            merchant=row.merchant,
            amount=row.amount,
            category=row.category,
            year=d.year,
            month=d.month,
            source=source,
        )
        db.add(txn)
        imported += 1
    db.commit()
    return {"imported": imported, "skipped_duplicates": skipped_duplicates}


# ---------------------------------------------------------------------------
# Budget auto-populate from historical averages
# ---------------------------------------------------------------------------

class AutoPopulateRequest(BaseModel):
    year: int
    month: int
    lookback_months: int = 3
    overwrite: bool = False


@app.post("/budget-targets/auto-populate")
def auto_populate_budget(body: AutoPopulateRequest, db: Session = Depends(get_db)):
    """Create budget targets for a month based on historical spending averages."""
    # Determine the N months before the target month
    target = body.year * 12 + body.month
    history_months = []
    for i in range(1, body.lookback_months + 1):
        val = target - i
        history_months.append((val // 12, val % 12 if val % 12 != 0 else 12))
        if val % 12 == 0:
            history_months[-1] = (val // 12 - 1, 12)

    # Sum per category per month in the lookback window
    cat_totals: dict[str, list[float]] = defaultdict(list)
    for yr, mo in history_months:
        rows = db.execute(
            select(models.Transaction.category, func.sum(models.Transaction.amount).label("total"))
            .where(models.Transaction.year == yr, models.Transaction.month == mo)
            .group_by(models.Transaction.category)
        ).all()
        for r in rows:
            cat_totals[r.category].append(r.total)

    created = 0
    skipped = 0
    for category, monthly_vals in cat_totals.items():
        avg = round(sum(monthly_vals) / len(monthly_vals), 2)
        existing = db.execute(
            select(models.BudgetTarget).where(
                models.BudgetTarget.category == category,
                models.BudgetTarget.year == body.year,
                models.BudgetTarget.month == body.month,
            )
        ).scalar_one_or_none()

        if existing:
            if body.overwrite:
                existing.amount = avg
                created += 1
            else:
                skipped += 1
        else:
            db.add(models.BudgetTarget(category=category, year=body.year, month=body.month, amount=avg))
            created += 1

    db.commit()
    return {"set": created, "skipped": skipped, "months_analyzed": len(history_months)}


class CopyBudgetRequest(BaseModel):
    from_year: int
    from_month: int
    to_year: int
    to_month: int
    overwrite: bool = False


@app.post("/budget-targets/copy-from-month")
def copy_budget_from_month(body: CopyBudgetRequest, db: Session = Depends(get_db)):
    """Copy budget targets from one month to another."""
    source = db.execute(
        select(models.BudgetTarget).where(
            models.BudgetTarget.year == body.from_year,
            models.BudgetTarget.month == body.from_month,
        )
    ).scalars().all()
    copied = 0
    skipped = 0
    for t in source:
        existing = db.execute(
            select(models.BudgetTarget).where(
                models.BudgetTarget.category == t.category,
                models.BudgetTarget.year == body.to_year,
                models.BudgetTarget.month == body.to_month,
            )
        ).scalar_one_or_none()
        if existing and not body.overwrite:
            skipped += 1
            continue
        if existing:
            existing.amount = t.amount
        else:
            db.add(models.BudgetTarget(
                category=t.category, year=body.to_year,
                month=body.to_month, amount=t.amount,
            ))
        copied += 1
    db.commit()
    return {"copied": copied, "skipped": skipped}


class RolloverRequest(BaseModel):
    year: int
    month: int
    carry_remainder: bool = True


@app.post("/budget-targets/rollover")
def rollover_budget(body: RolloverRequest, db: Session = Depends(get_db)):
    """Copy prior month's budget targets to current month. Optionally add unspent amount."""
    prior_month = body.month - 1 if body.month > 1 else 12
    prior_year = body.year if body.month > 1 else body.year - 1

    prior_targets = db.execute(
        select(models.BudgetTarget).where(
            models.BudgetTarget.year == prior_year,
            models.BudgetTarget.month == prior_month,
        )
    ).scalars().all()

    # Prior month actuals per category
    prior_actuals = {}
    if body.carry_remainder:
        rows = db.execute(
            select(models.Transaction.category, func.sum(models.Transaction.amount).label("total"))
            .where(models.Transaction.year == prior_year, models.Transaction.month == prior_month)
            .group_by(models.Transaction.category)
        ).all()
        prior_actuals = {r.category: r.total for r in rows}

    existing_cats = {t.category for t in db.execute(
        select(models.BudgetTarget).where(
            models.BudgetTarget.year == body.year,
            models.BudgetTarget.month == body.month,
        )
    ).scalars().all()}

    copied = 0
    for t in prior_targets:
        if t.category in existing_cats:
            continue
        rollover_amount = t.amount
        if body.carry_remainder:
            actual = prior_actuals.get(t.category, 0)
            unspent = t.amount - actual
            if unspent > 0:
                rollover_amount = t.amount + unspent
        db.add(models.BudgetTarget(category=t.category, year=body.year, month=body.month, amount=round(rollover_amount, 2)))
        copied += 1

    db.commit()
    return {"copied": copied}


@app.get("/transactions/anomalies")
def transaction_anomalies(year: int, month: int, threshold: float = 2.0, db: Session = Depends(get_db)):
    """Return transactions with amount > threshold × their category's per-transaction average."""
    # Compute per-category average transaction amount from prior 3 months
    target = year * 12 + month
    prior_avgs: dict[str, list[float]] = defaultdict(list)
    for i in range(1, 4):
        val = target - i
        py = (val - 1) // 12 + 1 if val % 12 == 0 else val // 12
        pm = 12 if val % 12 == 0 else val % 12
        rows = db.execute(
            select(models.Transaction.category, func.avg(models.Transaction.amount).label("avg"))
            .where(models.Transaction.year == py, models.Transaction.month == pm)
            .group_by(models.Transaction.category)
        ).all()
        for r in rows:
            prior_avgs[r.category].append(r.avg)

    avg_per_cat = {cat: sum(v) / len(v) for cat, v in prior_avgs.items() if v}

    txns = db.execute(
        select(models.Transaction).where(
            models.Transaction.year == year,
            models.Transaction.month == month,
        )
    ).scalars().all()

    anomalies = []
    for txn in txns:
        avg = avg_per_cat.get(txn.category)
        if avg and avg > 5 and txn.amount > avg * threshold:
            anomalies.append({
                "id": txn.id,
                "category": txn.category,
                "amount": txn.amount,
                "avg": round(avg, 2),
                "ratio": round(txn.amount / avg, 1),
            })
    return anomalies


class SplitRequest(BaseModel):
    amount_a: float
    category_a: str
    amount_b: float
    category_b: str


@app.post("/transactions/{txn_id}/split")
def split_transaction(txn_id: int, body: SplitRequest, db: Session = Depends(get_db)):
    """Split a transaction into two separate transactions."""
    txn = db.get(models.Transaction, txn_id)
    if not txn:
        raise HTTPException(404, "Transaction not found")
    if abs(body.amount_a + body.amount_b - txn.amount) > 0.02:
        raise HTTPException(400, f"Split amounts ({body.amount_a} + {body.amount_b}) must equal original ({txn.amount})")
    t1 = models.Transaction(
        date=txn.date, merchant=txn.merchant, amount=round(body.amount_a, 2),
        category=body.category_a, year=txn.year, month=txn.month,
        is_fixed=txn.is_fixed, is_recurring=txn.is_recurring, notes=txn.notes, source=txn.source
    )
    t2 = models.Transaction(
        date=txn.date, merchant=txn.merchant, amount=round(body.amount_b, 2),
        category=body.category_b, year=txn.year, month=txn.month,
        is_fixed=txn.is_fixed, is_recurring=txn.is_recurring, notes=txn.notes, source=txn.source
    )
    db.add(t1)
    db.add(t2)
    db.delete(txn)
    db.commit()
    return {"ok": True}


@app.get("/transactions/recurring-suggestions")
def recurring_suggestions(min_months: int = 3, db: Session = Depends(get_db)):
    """Find merchants appearing in 3+ different months at a consistent amount — likely recurring."""
    rows = db.execute(
        select(
            models.Transaction.merchant,
            models.Transaction.category,
            func.count(func.distinct(
                func.cast(models.Transaction.year * 100 + models.Transaction.month, models.Transaction.id.type)
            )).label("month_count"),
            func.avg(models.Transaction.amount).label("avg_amount"),
            func.min(models.Transaction.amount).label("min_amount"),
            func.max(models.Transaction.amount).label("max_amount"),
        )
        .group_by(models.Transaction.merchant, models.Transaction.category)
    ).all()

    suggestions = []
    for r in rows:
        month_count = r.month_count
        if month_count < min_months:
            continue
        # Consistent = max within 20% of min
        if r.min_amount > 0 and r.max_amount <= r.min_amount * 1.2:
            suggestions.append({
                "merchant": r.merchant,
                "category": r.category,
                "month_count": month_count,
                "avg_amount": round(r.avg_amount, 2),
                "is_consistent": True,
            })

    return sorted(suggestions, key=lambda x: -x["month_count"])


@app.get("/summary/multi-category-trend")
def multi_category_trend(categories: str, db: Session = Depends(get_db)):
    """Monthly spending for multiple categories. categories = comma-separated."""
    cat_list = [c.strip() for c in categories.split(",") if c.strip()]
    if not cat_list:
        return []

    rows = db.execute(
        select(
            models.Transaction.category,
            models.Transaction.year,
            models.Transaction.month,
            func.sum(models.Transaction.amount).label("total"),
        )
        .where(models.Transaction.category.in_(cat_list))
        .group_by(models.Transaction.category, models.Transaction.year, models.Transaction.month)
        .order_by(models.Transaction.year, models.Transaction.month)
    ).all()

    return [{"category": r.category, "year": r.year, "month": r.month, "total": round(r.total, 2)} for r in rows]


# ---------------------------------------------------------------------------
# Budget templates
# ---------------------------------------------------------------------------

class SaveTemplateRequest(BaseModel):
    name: str
    year: int
    month: int


class ApplyTemplateRequest(BaseModel):
    name: str
    year: int
    month: int
    overwrite: bool = False


@app.get("/budget-templates")
def list_budget_templates(db: Session = Depends(get_db)):
    rows = db.execute(
        select(models.AppSettings).where(
            models.AppSettings.key.like("budget_template:%")
        )
    ).scalars().all()
    templates = []
    for r in rows:
        name = r.key[len("budget_template:"):]
        try:
            items = json.loads(r.value)
        except Exception:
            items = []
        templates.append({"name": name, "count": len(items)})
    return templates


@app.post("/budget-templates/save")
def save_budget_template(body: SaveTemplateRequest, db: Session = Depends(get_db)):
    targets = db.execute(
        select(models.BudgetTarget).where(
            models.BudgetTarget.year == body.year,
            models.BudgetTarget.month == body.month,
        )
    ).scalars().all()
    items = [{"category": t.category, "amount": t.amount} for t in targets]
    key = f"budget_template:{body.name}"
    existing = db.get(models.AppSettings, key)
    if existing:
        existing.value = json.dumps(items)
    else:
        db.add(models.AppSettings(key=key, value=json.dumps(items)))
    db.commit()
    return {"name": body.name, "count": len(items)}


@app.post("/budget-templates/apply")
def apply_budget_template(body: ApplyTemplateRequest, db: Session = Depends(get_db)):
    key = f"budget_template:{body.name}"
    setting = db.get(models.AppSettings, key)
    if not setting:
        raise HTTPException(404, "Template not found")
    items = json.loads(setting.value)
    existing = {t.category: t for t in db.execute(
        select(models.BudgetTarget).where(
            models.BudgetTarget.year == body.year,
            models.BudgetTarget.month == body.month,
        )
    ).scalars().all()}
    applied = 0
    for item in items:
        if item["category"] in existing:
            if body.overwrite:
                existing[item["category"]].amount = item["amount"]
                applied += 1
        else:
            db.add(models.BudgetTarget(category=item["category"], year=body.year, month=body.month, amount=item["amount"]))
            applied += 1
    db.commit()
    return {"applied": applied}


@app.delete("/budget-templates/{name}", status_code=204)
def delete_budget_template(name: str, db: Session = Depends(get_db)):
    key = f"budget_template:{name}"
    setting = db.get(models.AppSettings, key)
    if not setting:
        raise HTTPException(404, "Template not found")
    db.delete(setting)
    db.commit()


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------

@app.post("/import")
async def import_file(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files are supported")

    content = await file.read()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        counts = import_xlsx(tmp_path, db)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Import failed: {type(e).__name__}: {e}")
    finally:
        os.unlink(tmp_path)

    return {
        "message": "Import complete",
        "transactions_imported": counts["transactions"],
        "income_imported": counts["income"],
        "records_updated": counts["skipped"],
    }


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

@app.get("/export")
def export_xlsx(year: int, month: Optional[int] = None, db: Session = Depends(get_db)):
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # Transactions sheet
    ws_txn = wb.active
    ws_txn.title = "Transactions"
    headers = ["Date", "Merchant", "Amount", "Category", "Month", "Year", "Source"]
    for col, h in enumerate(headers, 1):
        cell = ws_txn.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.font = Font(bold=True, color="FFFFFF")

    q = select(models.Transaction).where(models.Transaction.year == year)
    if month:
        q = q.where(models.Transaction.month == month)
    q = q.order_by(models.Transaction.date)
    txns = db.execute(q).scalars().all()

    for row_idx, txn in enumerate(txns, 2):
        ws_txn.cell(row=row_idx, column=1, value=txn.date)
        ws_txn.cell(row=row_idx, column=2, value=txn.merchant)
        ws_txn.cell(row=row_idx, column=3, value=txn.amount)
        ws_txn.cell(row=row_idx, column=4, value=txn.category)
        ws_txn.cell(row=row_idx, column=5, value=txn.month)
        ws_txn.cell(row=row_idx, column=6, value=txn.year)
        ws_txn.cell(row=row_idx, column=7, value=txn.source)

    # Monthly Summary sheet
    ws_sum = wb.create_sheet("Monthly Summary")
    sum_headers = ["Month", "Income", "Expenses", "Balance"]
    for col, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    monthly = monthly_summary(year, db)
    for row_idx, m in enumerate(monthly, 2):
        month_names = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
                       "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        ws_sum.cell(row=row_idx, column=1, value=month_names[m["month"]])
        ws_sum.cell(row=row_idx, column=2, value=m["income"])
        ws_sum.cell(row=row_idx, column=3, value=m["expenses"])
        ws_sum.cell(row=row_idx, column=4, value=m["balance"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"budget_{year}.xlsx" if not month else f"budget_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ---------------------------------------------------------------------------
# AI Insights
# ---------------------------------------------------------------------------

MONTH_NAMES = ["", "January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]


def _build_debt_context(db: Session) -> list[str]:
    """Build debt section for AI context, including payoff recommendations request."""
    debts = db.execute(select(models.Debt)).scalars().all()
    if not debts:
        return []

    lines = ["", "### Current Debts"]
    total_balance = 0
    total_min_payment = 0

    for d in debts:
        total_balance += d.current_balance
        total_min_payment += d.monthly_payment + d.monthly_extra
        debt_type = "Line of Credit" if d.debt_type == "loc" else "Loan"
        rate_str = f" @ {d.interest_rate * 100:.2f}% p.a." if d.interest_rate else " (0% interest)"

        if d.debt_type == "loc":
            available = max(0, (d.credit_limit or 0) - d.current_balance)
            monthly_interest = d.current_balance * (d.interest_rate / 12) if d.interest_rate else 0
            lines.append(
                f"- **{d.name}** ({debt_type}, {d.creditor}){rate_str}: "
                f"${d.current_balance:,.0f} outstanding / ${d.credit_limit:,.0f} limit "
                f"(${available:,.0f} available) — min payment ${d.monthly_payment:,.0f}/mo, "
                f"monthly interest ~${monthly_interest:,.0f}"
            )
        else:
            months_left = None
            if d.monthly_payment + d.monthly_extra > 0 and d.current_balance > 0:
                total_pmt = d.monthly_payment + d.monthly_extra
                if d.interest_rate:
                    r = d.interest_rate / 12
                    if total_pmt > d.current_balance * r:
                        months_left = int(math.ceil(math.log(total_pmt / (total_pmt - d.current_balance * r)) / math.log(1 + r)))
                else:
                    months_left = int(math.ceil(d.current_balance / total_pmt)) if total_pmt > 0 else None
            payoff_str = f", payoff in ~{months_left} months" if months_left else ""
            lines.append(
                f"- **{d.name}** ({debt_type}, {d.creditor}){rate_str}: "
                f"${d.current_balance:,.0f} remaining — "
                f"${d.monthly_payment + d.monthly_extra:,.0f}/mo total payment{payoff_str}"
            )

    lines += [
        f"- **Total debt: ${total_balance:,.0f}** | Total committed payments: ${total_min_payment:,.0f}/mo",
    ]
    return lines


def _build_insights_context(year: int, month: Optional[int], db: Session) -> str:
    """Build the AI prompt context. month=None means full-year annual analysis."""

    if month:
        return _build_monthly_context(year, month, db)
    return _build_annual_context(year, db)


def _build_monthly_context(year: int, month: int, db: Session) -> str:
    """Single-month analysis context."""
    cat_rows = db.execute(
        select(models.Transaction.category, func.sum(models.Transaction.amount).label("total"))
        .where(models.Transaction.year == year, models.Transaction.month == month)
        .group_by(models.Transaction.category)
        .order_by(func.sum(models.Transaction.amount).desc())
    ).all()

    income_rows = db.execute(
        select(models.Income.person, models.Income.income_type, models.Income.amount)
        .where(models.Income.year == year, models.Income.month == month)
        .order_by(models.Income.person)
    ).all()
    total_income = sum(r.amount for r in income_rows)

    target_rows = db.execute(
        select(models.BudgetTarget.category, models.BudgetTarget.amount)
        .where(models.BudgetTarget.year == year, models.BudgetTarget.month == month)
    ).all()
    targets = {r.category: r.amount for r in target_rows}

    # Historical averages excluding current month
    hist_totals = defaultdict(list)
    for r in db.execute(
        select(models.Transaction.year, models.Transaction.month, models.Transaction.category,
               func.sum(models.Transaction.amount).label("total"))
        .where(~((models.Transaction.year == year) & (models.Transaction.month == month)))
        .group_by(models.Transaction.year, models.Transaction.month, models.Transaction.category)
    ).all():
        hist_totals[r.category].append(r.total)
    hist_avg = {cat: sum(v) / len(v) for cat, v in hist_totals.items()}

    ytd_exp = db.execute(select(func.sum(models.Transaction.amount))
        .where(models.Transaction.year == year, models.Transaction.month <= month)).scalar() or 0
    ytd_inc = db.execute(select(func.sum(models.Income.amount))
        .where(models.Income.year == year, models.Income.month <= month)).scalar() or 0

    total_expenses = sum(r.total for r in cat_rows)
    savings = total_income - total_expenses
    savings_rate = (savings / total_income * 100) if total_income else 0
    month_label = MONTH_NAMES[month]

    lines = [
        "You are a personal finance advisor analyzing a Canadian household budget.",
        "",
        f"## Period: {month_label} {year}",
        "",
        "### Income",
    ]
    if income_rows:
        for r in income_rows:
            lines.append(f"- {r.person} ({r.income_type}): ${r.amount:,.0f}")
    else:
        lines.append("- No income recorded for this month")
    lines.append(f"- **Total household income: ${total_income:,.0f}**")
    lines += [
        "",
        f"### Spending by Category (total: ${total_expenses:,.0f})",
        "| Category | This Month | Budget | Hist. Avg | vs Budget | vs Avg |",
        "|---|---|---|---|---|---|",
    ]
    for r in cat_rows:
        budget = targets.get(r.category)
        avg = hist_avg.get(r.category)
        vs_budget = f"+${r.total - budget:,.0f} over" if budget and r.total > budget else (f"${budget - r.total:,.0f} under" if budget else "N/A")
        vs_avg = f"+${r.total - avg:,.0f} ({((r.total / avg) - 1) * 100:.0f}%)" if avg else "N/A"
        lines.append(f"| {r.category} | ${r.total:,.0f} | {'$' + f'{budget:,.0f}' if budget else 'N/A'} | {'$' + f'{avg:,.0f}' if avg else 'N/A'} | {vs_budget} | {vs_avg} |")

    lines += [
        "",
        "### Month Summary",
        f"- Net savings: ${savings:,.0f} ({savings_rate:.1f}% savings rate)",
        f"- YTD income: ${ytd_inc:,.0f} | YTD expenses: ${ytd_exp:,.0f} | YTD balance: ${ytd_inc - ytd_exp:,.0f}",
        "",
        "---",
        "",
        "Please provide actionable, specific financial insights for this household. Include:",
        "1. **Top spending concerns** — categories that are high vs budget or historical average",
        "2. **Positive patterns** — where they are doing well",
        "3. **Concrete suggestions** — specific ways to reduce spending with realistic targets",
        "4. **Savings outlook** — comment on the savings rate and any recommendations",
        "5. **One priority action** — the single most impactful thing they could do this month",
        "",
        "Keep the tone practical and encouraging. Use Canadian dollar amounts. Be specific with numbers.",
    ]

    debt_lines = _build_debt_context(db)
    if debt_lines:
        lines += debt_lines
        lines += [
            "",
            "---",
            "",
            "**Debt Payoff Recommendations:**",
            f"Given the household income of ${total_income:,.0f}/month and expenses of ${total_expenses:,.0f}/month (leaving ${savings:,.0f}/month):",
            "6. **Recommended monthly payment per debt** — calculate the optimal payment for each debt to pay them off efficiently. Show the math: how much goes to each debt, what order, and the projected payoff date.",
            "7. **Acceleration opportunity** — if they freed up $200-500/month, how much faster could they be debt-free?",
            "8. **Interest cost warning** — for any debts with interest rates, show the total interest they'll pay at current payment pace vs an accelerated pace.",
            "",
            "Keep the tone practical and encouraging. Use Canadian dollar amounts. Be specific with numbers.",
        ]

    return "\n".join(lines)


def _build_annual_context(year: int, db: Session) -> str:
    """Full-year analysis context."""
    # All transactions this year by category
    cat_rows = db.execute(
        select(models.Transaction.category, func.sum(models.Transaction.amount).label("total"))
        .where(models.Transaction.year == year)
        .group_by(models.Transaction.category)
        .order_by(func.sum(models.Transaction.amount).desc())
    ).all()

    # Month-by-month breakdown
    monthly_exp = db.execute(
        select(models.Transaction.month, func.sum(models.Transaction.amount).label("total"))
        .where(models.Transaction.year == year)
        .group_by(models.Transaction.month)
        .order_by(models.Transaction.month)
    ).all()
    monthly_inc = db.execute(
        select(models.Income.month, func.sum(models.Income.amount).label("total"))
        .where(models.Income.year == year)
        .group_by(models.Income.month)
        .order_by(models.Income.month)
    ).all()
    inc_by_month = {r.month: r.total for r in monthly_inc}
    exp_by_month = {r.month: r.total for r in monthly_exp}

    # Annual totals
    total_expenses = sum(r.total for r in cat_rows)
    total_income = db.execute(
        select(func.sum(models.Income.amount)).where(models.Income.year == year)
    ).scalar() or 0

    # Prior year for comparison
    prior_cat = db.execute(
        select(models.Transaction.category, func.sum(models.Transaction.amount).label("total"))
        .where(models.Transaction.year == year - 1)
        .group_by(models.Transaction.category)
    ).all()
    prior_totals = {r.category: r.total for r in prior_cat}
    prior_income = db.execute(
        select(func.sum(models.Income.amount)).where(models.Income.year == year - 1)
    ).scalar() or 0
    prior_expenses = sum(r.total for r in prior_cat)

    # Budget targets for the year (use any month's targets as reference — average them)
    target_rows = db.execute(
        select(models.BudgetTarget.category, func.avg(models.BudgetTarget.amount).label("avg_amount"))
        .where(models.BudgetTarget.year == year)
        .group_by(models.BudgetTarget.category)
    ).all()
    targets = {r.category: r.avg_amount * 12 for r in target_rows}  # annualise monthly budgets

    # Peak spending month per top category
    peak_rows = db.execute(
        select(models.Transaction.category, models.Transaction.month,
               func.sum(models.Transaction.amount).label("total"))
        .where(models.Transaction.year == year)
        .group_by(models.Transaction.category, models.Transaction.month)
    ).all()
    cat_monthly = defaultdict(dict)
    for r in peak_rows:
        cat_monthly[r.category][r.month] = r.total

    # Savings by month
    savings_by_month = {m: inc_by_month.get(m, 0) - exp_by_month.get(m, 0) for m in range(1, 13)}
    best_month = max(savings_by_month, key=savings_by_month.get)
    worst_month = min(savings_by_month, key=savings_by_month.get)

    total_savings = total_income - total_expenses
    savings_rate = (total_savings / total_income * 100) if total_income else 0

    months_with_data = [m for m in range(1, 13) if exp_by_month.get(m, 0) > 0]
    last_month_label = MONTH_NAMES[max(months_with_data)] if months_with_data else "N/A"

    lines = [
        "You are a personal finance advisor analyzing a Canadian household budget.",
        "",
        f"## Period: Full Year {year} (through {last_month_label})",
        "",
        f"### Annual Totals",
        f"- Total household income: ${total_income:,.0f}",
        f"- Total expenses: ${total_expenses:,.0f}",
        f"- Net savings: **${total_savings:,.0f}** ({savings_rate:.1f}% savings rate)",
    ]

    if prior_income or prior_expenses:
        prior_savings = prior_income - prior_expenses
        inc_chg = ((total_income - prior_income) / prior_income * 100) if prior_income else 0
        exp_chg = ((total_expenses - prior_expenses) / prior_expenses * 100) if prior_expenses else 0
        lines += [
            "",
            f"### Year-over-Year vs {year - 1}",
            f"- Income: ${total_income:,.0f} vs ${prior_income:,.0f} ({inc_chg:+.1f}%)",
            f"- Expenses: ${total_expenses:,.0f} vs ${prior_expenses:,.0f} ({exp_chg:+.1f}%)",
            f"- Savings: ${total_savings:,.0f} vs ${prior_savings:,.0f}",
        ]

    lines += [
        "",
        "### Month-by-Month Summary",
        "| Month | Income | Expenses | Savings | Rate |",
        "|---|---|---|---|---|",
    ]
    for m in range(1, 13):
        inc = inc_by_month.get(m, 0)
        exp = exp_by_month.get(m, 0)
        if inc == 0 and exp == 0:
            continue
        sav = inc - exp
        rate = f"{sav / inc * 100:.0f}%" if inc else "—"
        lines.append(f"| {MONTH_NAMES[m][:3]} | ${inc:,.0f} | ${exp:,.0f} | ${sav:,.0f} | {rate} |")

    lines += [
        "",
        f"### Spending by Category (annual total: ${total_expenses:,.0f})",
        "| Category | Annual Total | Annual Budget | vs {yr_prior} | Peak Month |".format(yr_prior=year - 1),
        "|---|---|---|---|---|",
    ]
    for r in cat_rows:
        budget = targets.get(r.category)
        prior = prior_totals.get(r.category)
        vs_prior = f"+${r.total - prior:,.0f} ({((r.total / prior) - 1) * 100:.0f}%)" if prior else "N/A"
        peak_m = max(cat_monthly.get(r.category, {1: 0}), key=cat_monthly.get(r.category, {1: 0}).get)
        peak_label = MONTH_NAMES[peak_m][:3] if cat_monthly.get(r.category) else "—"
        vs_budget = f"${budget - r.total:,.0f} under" if budget and r.total <= budget else (f"+${r.total - budget:,.0f} over" if budget else "N/A")
        lines.append(f"| {r.category} | ${r.total:,.0f} | {'$' + f'{budget:,.0f}' if budget else 'N/A'} ({vs_budget}) | {vs_prior} | {peak_label} |")

    lines += [
        "",
        f"### Savings Patterns",
        f"- Best month: {MONTH_NAMES[best_month]} (saved ${savings_by_month[best_month]:,.0f})",
        f"- Worst month: {MONTH_NAMES[worst_month]} (saved ${savings_by_month[worst_month]:,.0f})",
        "",
        "---",
        "",
        "Please provide a comprehensive annual financial review for this household. Include:",
        "1. **Annual performance summary** — overall savings rate, income vs expenses trend vs prior year",
        "2. **Top spending categories** — which categories drove the most spending and how they compare to prior year",
        "3. **Budget adherence** — where they stayed within budget and where they overspent",
        "4. **Seasonal patterns** — months with unusually high/low spending and why that might be",
        "5. **Savings rate analysis** — is the rate healthy? What would move it meaningfully?",
        "6. **Top 3 priorities for next year** — specific, actionable goals based on this year's data",
        "",
        "Keep the tone practical and encouraging. Use Canadian dollar amounts. Be specific with numbers.",
    ]

    debt_lines = _build_debt_context(db)
    if debt_lines:
        avg_monthly_income = total_income / len([m for m in range(1, 13) if inc_by_month.get(m, 0) > 0]) if total_income else 0
        avg_monthly_expenses = total_expenses / len([m for m in range(1, 13) if exp_by_month.get(m, 0) > 0]) if total_expenses else 0
        avg_monthly_savings = avg_monthly_income - avg_monthly_expenses
        lines += debt_lines
        lines += [
            "",
            "---",
            "",
            "**Debt Payoff Recommendations:**",
            f"Average monthly income: ${avg_monthly_income:,.0f} | Average monthly expenses: ${avg_monthly_expenses:,.0f} | Average monthly surplus: ${avg_monthly_savings:,.0f}",
            "7. **Recommended monthly payment per debt** — based on their income and spending patterns, calculate the optimal payment for each debt. Show the math: recommended amount per debt, payoff order, and projected payoff dates.",
            "8. **Optimal payoff strategy** — avalanche (highest interest first) vs snowball (lowest balance first). Given their specific debts, which saves more money?",
            "9. **Acceleration scenario** — if they put an extra $300/month toward debt, which debt should receive it first and how much sooner would they be debt-free?",
            "10. **Total interest cost** — how much interest will they pay at current pace? How much would they save with the accelerated plan?",
            "",
            "Keep the tone practical and encouraging. Use Canadian dollar amounts. Be specific with numbers.",
        ]

    return "\n".join(lines)


@app.get("/insights")
async def get_insights(year: int, month: Optional[int] = None, db: Session = Depends(get_db)):
    import anthropic as anthropic_sdk

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise HTTPException(400, "ANTHROPIC_API_KEY environment variable is not set")

    context = _build_insights_context(year, month, db)

    async def stream_insights():
        client = anthropic_sdk.Anthropic(api_key=api_key)
        try:
            with client.messages.stream(
                model="claude-opus-4-6",
                max_tokens=8192,
                thinking={"type": "enabled", "budget_tokens": 5000},
                messages=[{"role": "user", "content": context}],
            ) as stream:
                for text in stream.text_stream:
                    yield f"data: {json.dumps({'text': text})}\n\n"
        except Exception as e:
            yield f"data: {json.dumps({'error': str(e)})}\n\n"
        yield "data: [DONE]\n\n"

    return StreamingResponse(stream_insights(), media_type="text/event-stream")


class InsightsLogCreate(BaseModel):
    year: int
    month: int  # 0 = annual
    content: str


@app.post("/insights/log", status_code=201)
def save_insights_log(body: InsightsLogCreate, db: Session = Depends(get_db)):
    entry = models.InsightsLog(
        year=body.year,
        month=body.month,
        generated_at=datetime.datetime.now(),
        content=body.content,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return {"id": entry.id, "generated_at": entry.generated_at.isoformat()}


@app.get("/insights/log")
def list_insights_log(db: Session = Depends(get_db)):
    rows = db.execute(
        select(models.InsightsLog).order_by(models.InsightsLog.generated_at.desc())
    ).scalars().all()
    return [
        {
            "id": r.id,
            "year": r.year,
            "month": r.month,
            "generated_at": r.generated_at.isoformat(),
            "content": r.content,
        }
        for r in rows
    ]


@app.delete("/insights/log/{entry_id}", status_code=204)
def delete_insights_log(entry_id: int, db: Session = Depends(get_db)):
    entry = db.get(models.InsightsLog, entry_id)
    if not entry:
        raise HTTPException(404, "Not found")
    db.delete(entry)
    db.commit()


# ── Net Worth Snapshots ─────────────────────────────────────────────────────

class NetWorthSnapshotOut(BaseModel):
    id: int
    snapshot_date: str
    total_assets: float
    liquid_assets: float
    illiquid_assets: float
    total_debts: float
    net_worth: float
    liquid_net_worth: float
    notes: Optional[str] = None

    class Config:
        from_attributes = True


@app.post("/net-worth/snapshot", response_model=NetWorthSnapshotOut, status_code=201)
def create_net_worth_snapshot(
    snapshot_notes: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Take a snapshot of current net worth computed from live assets/debts."""
    assets = db.execute(select(models.Asset)).scalars().all()
    debts = db.execute(select(models.Debt)).scalars().all()

    total_assets = sum(a.balance for a in assets)
    liquid_assets = sum(a.balance for a in assets if a.liquidity == "liquid")
    illiquid_assets = sum(a.balance for a in assets if a.liquidity == "illiquid")
    total_debts = sum(d.current_balance for d in debts)
    non_mortgage_debts = sum(d.current_balance for d in debts if d.debt_type != "mortgage")

    snap = models.NetWorthSnapshot(
        snapshot_date=datetime.date.today().isoformat(),
        total_assets=total_assets,
        liquid_assets=liquid_assets,
        illiquid_assets=illiquid_assets,
        total_debts=total_debts,
        net_worth=total_assets - total_debts,
        liquid_net_worth=liquid_assets - non_mortgage_debts,
        notes=snapshot_notes,
    )
    db.add(snap)
    db.commit()
    db.refresh(snap)
    return snap


@app.get("/net-worth/history", response_model=list[NetWorthSnapshotOut])
def get_net_worth_history(db: Session = Depends(get_db)):
    rows = db.execute(
        select(models.NetWorthSnapshot).order_by(models.NetWorthSnapshot.snapshot_date)
    ).scalars().all()
    return rows


@app.delete("/net-worth/snapshot/{snap_id}", status_code=204)
def delete_net_worth_snapshot(snap_id: int, db: Session = Depends(get_db)):
    snap = db.get(models.NetWorthSnapshot, snap_id)
    if not snap:
        raise HTTPException(404, "Snapshot not found")
    db.delete(snap)
    db.commit()


# ── Savings Goals ───────────────────────────────────────────────────────────

class SavingsGoalCreate(BaseModel):
    name: str
    target_amount: float
    current_amount: float = 0.0
    target_date: Optional[str] = None
    linked_asset_id: Optional[int] = None
    notes: Optional[str] = None


class SavingsGoalOut(BaseModel):
    id: int
    name: str
    target_amount: float
    current_amount: float
    target_date: Optional[str] = None
    linked_asset_id: Optional[int] = None
    notes: Optional[str] = None
    progress_pct: float = 0.0

    class Config:
        from_attributes = True


@app.get("/savings-goals", response_model=list[SavingsGoalOut])
def list_savings_goals(db: Session = Depends(get_db)):
    goals = db.execute(select(models.SavingsGoal)).scalars().all()
    result = []
    for g in goals:
        current = g.current_amount
        # If linked to an asset, use the asset's live balance
        if g.linked_asset_id:
            asset = db.get(models.Asset, g.linked_asset_id)
            if asset:
                current = asset.balance
        pct = round((current / g.target_amount) * 100, 1) if g.target_amount > 0 else 0.0
        result.append(SavingsGoalOut(
            id=g.id,
            name=g.name,
            target_amount=g.target_amount,
            current_amount=current,
            target_date=g.target_date,
            linked_asset_id=g.linked_asset_id,
            notes=g.notes,
            progress_pct=pct,
        ))
    return result


@app.post("/savings-goals", response_model=SavingsGoalOut, status_code=201)
def create_savings_goal(body: SavingsGoalCreate, db: Session = Depends(get_db)):
    goal = models.SavingsGoal(**body.model_dump())
    db.add(goal)
    db.commit()
    db.refresh(goal)
    pct = round((goal.current_amount / goal.target_amount) * 100, 1) if goal.target_amount > 0 else 0.0
    return SavingsGoalOut(**{c.name: getattr(goal, c.name) for c in models.SavingsGoal.__table__.columns}, progress_pct=pct)


@app.put("/savings-goals/{goal_id}", response_model=SavingsGoalOut)
def update_savings_goal(goal_id: int, body: SavingsGoalCreate, db: Session = Depends(get_db)):
    goal = db.get(models.SavingsGoal, goal_id)
    if not goal:
        raise HTTPException(404, "Goal not found")
    for k, v in body.model_dump().items():
        setattr(goal, k, v)
    db.commit()
    db.refresh(goal)
    pct = round((goal.current_amount / goal.target_amount) * 100, 1) if goal.target_amount > 0 else 0.0
    return SavingsGoalOut(**{c.name: getattr(goal, c.name) for c in models.SavingsGoal.__table__.columns}, progress_pct=pct)


@app.delete("/savings-goals/{goal_id}", status_code=204)
def delete_savings_goal(goal_id: int, db: Session = Depends(get_db)):
    goal = db.get(models.SavingsGoal, goal_id)
    if not goal:
        raise HTTPException(404, "Goal not found")
    db.delete(goal)
    db.commit()


# ── Tax Summary ─────────────────────────────────────────────────────────────

@app.get("/tax-summary")
def get_tax_summary(year: int, db: Session = Depends(get_db)):
    """Canadian year-end tax summary: income by person, RRSP, ESPP, donations."""
    income_rows = db.execute(
        select(models.Income).where(models.Income.year == year)
    ).scalars().all()

    # Group income by person
    by_person: dict = {}
    for row in income_rows:
        p = row.person
        if p not in by_person:
            by_person[p] = {
                "gross_income": 0.0,
                "rrsp_employee": 0.0,
                "rrsp_employer": 0.0,
                "espp_deduction": 0.0,
                "other_income": 0.0,
            }
        if row.income_type in ("base", "commission"):
            by_person[p]["gross_income"] += row.amount
            by_person[p]["rrsp_employee"] += row.rrsp_employee or 0.0
            by_person[p]["rrsp_employer"] += row.rrsp_employer or 0.0
            by_person[p]["espp_deduction"] += row.espp_deduction or 0.0
        else:
            by_person[p]["other_income"] += row.amount

    # Charitable donations from transactions
    donation_rows = db.execute(
        select(models.Transaction).where(
            models.Transaction.year == year,
            models.Transaction.category == "Charitable Donations",
        )
    ).scalars().all()
    total_donations = sum(t.amount for t in donation_rows)

    # ESPP purchases for the year
    espp_purchases = db.execute(
        select(models.EsppPurchase).where(
            models.EsppPurchase.purchase_date.like(f"{year}%")
        )
    ).scalars().all()
    espp_summary = [
        {
            "purchase_date": e.purchase_date,
            "total_deducted": e.total_deducted,
            "shares_purchased": e.shares_purchased,
            "purchase_price": e.purchase_price,
            "market_price": e.market_price,
            "discount_benefit": round((e.market_price - e.purchase_price) * e.shares_purchased, 2) if e.shares_purchased else 0.0,
        }
        for e in espp_purchases
    ]

    return {
        "year": year,
        "by_person": by_person,
        "total_donations": total_donations,
        "espp_purchases": espp_summary,
        "total_rrsp_employee": sum(p["rrsp_employee"] for p in by_person.values()),
        "total_rrsp_employer": sum(p["rrsp_employer"] for p in by_person.values()),
        "total_espp_deduction": sum(p["espp_deduction"] for p in by_person.values()),
    }


# ── Net Worth Forecast ──────────────────────────────────────────────────────

@app.get("/forecast/networth")
def forecast_networth(months: int = Query(default=12, ge=1, le=60), db: Session = Depends(get_db)):
    """12-month net worth projection based on recent average monthly surplus."""
    # Get current net worth
    assets = db.execute(select(models.Asset)).scalars().all()
    debts = db.execute(select(models.Debt)).scalars().all()
    current_nw = sum(a.balance for a in assets) - sum(d.current_balance for d in debts)

    # Compute avg monthly surplus from last 6 months of income vs spending
    today = datetime.date.today()
    results = []
    monthly_surpluses = []
    for i in range(6, 0, -1):
        d = today - datetime.timedelta(days=30 * i)
        y, m = d.year, d.month
        income_total = db.execute(
            select(func.sum(models.Income.amount)).where(
                models.Income.year == y, models.Income.month == m
            )
        ).scalar() or 0.0
        spending_total = db.execute(
            select(func.sum(models.Transaction.amount)).where(
                models.Transaction.year == y,
                models.Transaction.month == m,
                models.Transaction.category != "Income",
            )
        ).scalar() or 0.0
        if income_total > 0:
            monthly_surpluses.append(income_total - spending_total)

    avg_monthly_surplus = sum(monthly_surpluses) / len(monthly_surpluses) if monthly_surpluses else 0.0

    # Project forward
    projected_nw = current_nw
    for i in range(1, months + 1):
        projected_nw += avg_monthly_surplus
        future_date = today + datetime.timedelta(days=30 * i)
        results.append({
            "month": future_date.strftime("%b %Y"),
            "projected_net_worth": round(projected_nw, 2),
        })

    return {
        "current_net_worth": round(current_nw, 2),
        "avg_monthly_surplus": round(avg_monthly_surplus, 2),
        "forecast": results,
    }


# ── Recurring Bills ─────────────────────────────────────────────────────────

class RecurringBillCreate(BaseModel):
    name: str
    merchant: str
    amount: float
    frequency: str = "monthly"
    due_day: Optional[int] = None
    category: Optional[str] = None
    last_seen: Optional[str] = None
    is_active: bool = True
    notes: Optional[str] = None


class RecurringBillOut(BaseModel):
    id: int
    name: str
    merchant: str
    amount: float
    frequency: str
    due_day: Optional[int] = None
    category: Optional[str] = None
    last_seen: Optional[str] = None
    is_active: bool
    notes: Optional[str] = None

    class Config:
        from_attributes = True


@app.get("/bills", response_model=list[RecurringBillOut])
def list_bills(db: Session = Depends(get_db)):
    return db.execute(
        select(models.RecurringBill).order_by(nullslast(models.RecurringBill.due_day), models.RecurringBill.name)
    ).scalars().all()


@app.post("/bills", response_model=RecurringBillOut, status_code=201)
def create_bill(body: RecurringBillCreate, db: Session = Depends(get_db)):
    bill = models.RecurringBill(**body.model_dump())
    db.add(bill)
    db.commit()
    db.refresh(bill)
    return bill


@app.put("/bills/{bill_id}", response_model=RecurringBillOut)
def update_bill(bill_id: int, body: RecurringBillCreate, db: Session = Depends(get_db)):
    bill = db.get(models.RecurringBill, bill_id)
    if not bill:
        raise HTTPException(404, "Bill not found")
    for k, v in body.model_dump().items():
        setattr(bill, k, v)
    db.commit()
    db.refresh(bill)
    return bill


@app.delete("/bills/{bill_id}", status_code=204)
def delete_bill(bill_id: int, db: Session = Depends(get_db)):
    bill = db.get(models.RecurringBill, bill_id)
    if not bill:
        raise HTTPException(404, "Bill not found")
    db.delete(bill)
    db.commit()


@app.get("/bills/upcoming")
def get_upcoming_bills(db: Session = Depends(get_db)):
    """Get bills with upcoming due dates this month and next month."""
    today = datetime.date.today()
    bills = db.execute(
        select(models.RecurringBill).where(models.RecurringBill.is_active == True)
    ).scalars().all()

    upcoming = []
    for bill in bills:
        if bill.due_day:
            # This month's due date
            try:
                import calendar
                last_day = calendar.monthrange(today.year, today.month)[1]
                due_day = min(bill.due_day, last_day)
                this_month_due = datetime.date(today.year, today.month, due_day)
                if this_month_due >= today:
                    days_until = (this_month_due - today).days
                    upcoming.append({
                        "id": bill.id,
                        "name": bill.name,
                        "merchant": bill.merchant,
                        "amount": bill.amount,
                        "due_date": this_month_due.isoformat(),
                        "days_until": days_until,
                        "frequency": bill.frequency,
                    })
                else:
                    # Next month
                    next_month = today.month + 1 if today.month < 12 else 1
                    next_year = today.year if today.month < 12 else today.year + 1
                    last_day_next = calendar.monthrange(next_year, next_month)[1]
                    due_day_next = min(bill.due_day, last_day_next)
                    next_due = datetime.date(next_year, next_month, due_day_next)
                    days_until = (next_due - today).days
                    upcoming.append({
                        "id": bill.id,
                        "name": bill.name,
                        "merchant": bill.merchant,
                        "amount": bill.amount,
                        "due_date": next_due.isoformat(),
                        "days_until": days_until,
                        "frequency": bill.frequency,
                    })
            except ValueError:
                pass

    upcoming.sort(key=lambda x: x["days_until"])
    return upcoming


# ── Net Worth Milestones ─────────────────────────────────────────────────────

class MilestoneCreate(BaseModel):
    label: str
    target_amount: float
    notes: Optional[str] = None

class MilestoneOut(BaseModel):
    id: int
    label: str
    target_amount: float
    achieved_at: Optional[str] = None
    notes: Optional[str] = None
    is_achieved: bool = False
    progress_pct: float = 0.0
    model_config = {"from_attributes": True}

@app.get("/net-worth/milestones", response_model=list[MilestoneOut])
def list_milestones(db: Session = Depends(get_db)):
    milestones = db.execute(
        select(models.NetWorthMilestone).order_by(models.NetWorthMilestone.target_amount)
    ).scalars().all()
    assets = db.execute(select(models.Asset)).scalars().all()
    debts = db.execute(select(models.Debt)).scalars().all()
    current_nw = sum(a.balance for a in assets) - sum(d.current_balance for d in debts)
    result = []
    for m in milestones:
        pct = round(min((current_nw / m.target_amount) * 100, 100), 1) if m.target_amount > 0 else 0.0
        is_achieved = current_nw >= m.target_amount
        # Auto-mark achieved_at if just crossed
        if is_achieved and not m.achieved_at:
            m.achieved_at = datetime.date.today().isoformat()
            db.commit()
        result.append(MilestoneOut(
            id=m.id, label=m.label, target_amount=m.target_amount,
            achieved_at=m.achieved_at, notes=m.notes,
            is_achieved=is_achieved, progress_pct=pct,
        ))
    return result

@app.post("/net-worth/milestones", response_model=MilestoneOut, status_code=201)
def create_milestone(body: MilestoneCreate, db: Session = Depends(get_db)):
    m = models.NetWorthMilestone(**body.model_dump())
    db.add(m)
    db.commit()
    db.refresh(m)
    return MilestoneOut(id=m.id, label=m.label, target_amount=m.target_amount,
                        achieved_at=m.achieved_at, notes=m.notes)

@app.delete("/net-worth/milestones/{mid}", status_code=204)
def delete_milestone(mid: int, db: Session = Depends(get_db)):
    m = db.get(models.NetWorthMilestone, mid)
    if not m:
        raise HTTPException(404, "Milestone not found")
    db.delete(m)
    db.commit()


# ── Merchant Category Rules ──────────────────────────────────────────────────

class MerchantRuleUpsert(BaseModel):
    merchant_pattern: str
    category: str

@app.post("/merchant-rules", status_code=200)
def upsert_merchant_rule(body: MerchantRuleUpsert, db: Session = Depends(get_db)):
    """Save or update a merchant→category rule (called after user corrects category)."""
    existing = db.execute(
        select(models.MerchantRule).where(models.MerchantRule.merchant_pattern == body.merchant_pattern)
    ).scalar_one_or_none()
    if existing:
        existing.category = body.category
        existing.updated_at = datetime.date.today().isoformat()
    else:
        rule = models.MerchantRule(
            merchant_pattern=body.merchant_pattern,
            category=body.category,
            updated_at=datetime.date.today().isoformat(),
        )
        db.add(rule)
    db.commit()
    return {"ok": True}

@app.get("/merchant-rules")
def list_merchant_rules(db: Session = Depends(get_db)):
    rules = db.execute(select(models.MerchantRule).order_by(models.MerchantRule.merchant_pattern)).scalars().all()
    return [{"id": r.id, "merchant_pattern": r.merchant_pattern, "category": r.category, "updated_at": r.updated_at} for r in rules]

@app.delete("/merchant-rules/{rule_id}", status_code=204)
def delete_merchant_rule(rule_id: int, db: Session = Depends(get_db)):
    rule = db.get(models.MerchantRule, rule_id)
    if not rule:
        raise HTTPException(404, "Rule not found")
    db.delete(rule)
    db.commit()


# ── Recurring Transaction Detection ─────────────────────────────────────────

@app.get("/transactions/missing-recurring")
def missing_recurring(db: Session = Depends(get_db)):
    """Find merchants that appeared in each of the last 3 months but not the current month."""
    today = datetime.date.today()
    current_year, current_month = today.year, today.month

    # Get the last 3 months
    check_months = []
    for i in range(1, 4):
        m = current_month - i
        y = current_year
        while m <= 0:
            m += 12
            y -= 1
        check_months.append((y, m))

    # Find merchants present in ALL 3 prior months
    merchant_months: dict = {}
    for y, m in check_months:
        rows = db.execute(
            select(models.Transaction.merchant, func.sum(models.Transaction.amount).label("total"))
            .where(models.Transaction.year == y, models.Transaction.month == m)
            .group_by(models.Transaction.merchant)
        ).all()
        for r in rows:
            if r.merchant not in merchant_months:
                merchant_months[r.merchant] = {"months": 0, "avg_amount": 0, "amounts": []}
            merchant_months[r.merchant]["months"] += 1
            merchant_months[r.merchant]["amounts"].append(r.total)

    # Keep only merchants that appeared in all 3 months
    recurring = {
        m: v for m, v in merchant_months.items()
        if v["months"] >= 3
    }

    # Check which are missing from current month
    current_merchants = set(
        r[0] for r in db.execute(
            select(models.Transaction.merchant)
            .where(models.Transaction.year == current_year, models.Transaction.month == current_month)
            .distinct()
        ).all()
    )

    missing = []
    for merchant, data in recurring.items():
        if merchant not in current_merchants:
            avg = round(sum(data["amounts"]) / len(data["amounts"]), 2)
            missing.append({
                "merchant": merchant,
                "avg_amount": avg,
                "months_seen": data["months"],
            })

    missing.sort(key=lambda x: x["avg_amount"], reverse=True)
    return missing


# ── PDF Bank Statement Parser ────────────────────────────────────────────────

@app.post("/parse-pdf")
async def parse_pdf(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Parse a bank statement PDF and extract transactions."""
    try:
        import pdfplumber
    except ImportError:
        raise HTTPException(500, "pdfplumber not installed")

    contents = await file.read()
    rows_out = []
    page_count = 0

    try:
        with pdfplumber.open(io.BytesIO(contents)) as pdf:
            page_count = len(pdf.pages)
            for page in pdf.pages:
                # Try table extraction first
                tables = page.extract_tables()
                for table in tables:
                    if not table or len(table) < 2:
                        continue
                    # Find header row
                    header = [str(c).lower().strip() if c else "" for c in table[0]]

                    def col_idx(names):
                        for name in names:
                            for i, h in enumerate(header):
                                if name in h:
                                    return i
                        return None

                    date_col = col_idx(["date", "transaction date", "posted"])
                    desc_col = col_idx(["description", "desc", "merchant", "details", "activity", "payee", "transaction"])
                    debit_col = col_idx(["debit", "withdrawal", "charges", "amount"])
                    credit_col = col_idx(["credit", "payment", "deposit"])
                    amt_col = col_idx(["amount"]) if debit_col is None else None

                    if date_col is None or desc_col is None:
                        continue

                    for row in table[1:]:
                        if not row or all(not c for c in row):
                            continue
                        date_str = str(row[date_col]).strip() if date_col < len(row) and row[date_col] else ""
                        merchant = str(row[desc_col]).strip() if desc_col < len(row) and row[desc_col] else ""
                        if not date_str or not merchant or merchant.lower() in ("", "none"):
                            continue

                        parsed_date = _parse_date(date_str)
                        if not parsed_date:
                            continue

                        amount = None
                        # Try debit column first
                        if debit_col is not None and debit_col < len(row) and row[debit_col]:
                            amount = _parse_amount(str(row[debit_col]))
                        # Fall back to amount column
                        if (amount is None or amount <= 0) and amt_col is not None and amt_col < len(row) and row[amt_col]:
                            val = _parse_amount(str(row[amt_col]))
                            if val and val > 0:
                                amount = val

                        if amount is None or amount <= 0:
                            continue

                        # Skip payment/credit rows
                        if credit_col is not None and credit_col < len(row) and row[credit_col]:
                            credit_val = _parse_amount(str(row[credit_col]))
                            if credit_val and credit_val > 0 and (amount is None or amount == 0):
                                continue

                        is_dup = db.execute(
                            select(models.Transaction).where(
                                models.Transaction.date == datetime.date.fromisoformat(parsed_date),
                                models.Transaction.merchant == merchant,
                                func.round(models.Transaction.amount, 2) == round(amount, 2),
                            )
                        ).first() is not None

                        suggested = _lookup_category(merchant, db)
                        rows_out.append({
                            "date": parsed_date,
                            "merchant": merchant,
                            "amount": round(amount, 2),
                            "suggested_category": suggested or "",
                            "confidence": "high" if suggested else "low",
                            "is_duplicate": is_dup,
                        })

                # If no tables found, try text-based extraction
                if not rows_out:
                    text = page.extract_text() or ""
                    lines = [l.strip() for l in text.splitlines() if l.strip()]
                    # Look for lines matching: date + description + amount pattern
                    for line in lines:
                        # Try to match: date at start, amount at end
                        m = re.match(
                            r'^(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}-\d{2}-\d{2}|\d{1,2}\s+\w+\s+\d{2,4})\s+(.+?)\s+(\$?[\d,]+\.\d{2})$',
                            line
                        )
                        if m:
                            parsed_date = _parse_date(m.group(1))
                            merchant = m.group(2).strip()
                            amount = _parse_amount(m.group(3))
                            if parsed_date and merchant and amount and amount > 0:
                                is_dup = db.execute(
                                    select(models.Transaction).where(
                                        models.Transaction.date == datetime.date.fromisoformat(parsed_date),
                                        models.Transaction.merchant == merchant,
                                        func.round(models.Transaction.amount, 2) == round(amount, 2),
                                    )
                                ).first() is not None
                                suggested = _lookup_category(merchant, db)
                                rows_out.append({
                                    "date": parsed_date,
                                    "merchant": merchant,
                                    "amount": round(amount, 2),
                                    "suggested_category": suggested or "",
                                    "confidence": "high" if suggested else "low",
                                    "is_duplicate": is_dup,
                                })
    except Exception as e:
        raise HTTPException(400, f"Could not parse PDF: {str(e)}")

    # Deduplicate rows_out by (date, merchant, amount)
    seen = set()
    unique_rows = []
    for r in rows_out:
        key = (r["date"], r["merchant"], r["amount"])
        if key not in seen:
            seen.add(key)
            unique_rows.append(r)

    return {"rows": unique_rows, "pages": page_count}
