"""
Handles importing budget data from FY24, FY25, and FY26 xlsx files.

FY24 format: sheets named "CC - Jan 24", "CB - Jan 24"
  CC sheets have two cards side by side: VISA (cols B-E) and AMEX (cols E-H)
  Headers at row 5: ITEM, DATE, CHARGE, CATEGORY

FY25/FY26 format: sheets named "Jan-2025 (CC)", "Jan - 2026 (CC)"
  Single card per sheet, headers at row 1: DATE, ITEM, AMOUNT, CATEGORY
"""
from __future__ import annotations
import re
from datetime import date, datetime
import openpyxl
from sqlalchemy.orm import Session
from models import Transaction, Income


MONTH_NAMES = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    "january": 1, "february": 2, "march": 3, "april": 4,
    "june": 6, "july": 7, "august": 8, "september": 9,
    "october": 10, "november": 11, "december": 12,
}

# Only exact 3-letter abbreviations — used for income row matching
MONTH_SHORT = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
               "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}

# Normalize category names to consistent Title Case
CATEGORY_MAP = {
    "HOUSEHOLD": "Home",
    "HEALTH": "Health & Beauty",
    "DELIVERY": "Take Out",
    "SUBSCRIPTIONS": "Entertainment",
    "DINE IN": "Dining",
    "TAKE OUT": "Take Out",
    "DINE-IN": "Dining",
    "TAKEOUT": "Take Out",
    "GROCERIES": "Groceries",
    "GAS": "Gas",
    "MOBILE": "Mobile",
    "INTERNET": "Internet",
    "ALCOHOL": "Alcohol",
    "CANNABIS": "Cannabis",
    "COFFEE": "Coffee",
    "ENTERTAINMENT": "Entertainment",
    "GIFTS": "Gifts",
    "CLOTHES": "Clothes",
    "HEALTH & BEAUTY": "Health & Beauty",
    "HOME": "Home",
    "HOME SECURITY": "Security",
    "TELUS SECURITY": "Security",
    "PET FOOD/TOYS": "Pet Food & Toys",
    "DAY CARE": "Day Care",
    "VET BILL": "Vet",
    "PET INSURANCE": "Pet Insurance",
    "CAR INSURANCE": "Car Insurance",
    "HOME INSURANCE": "Home Insurance",
    "CAR PAYMENT": "Car Payment",
    "CAR": "Car",
    "MORTGAGE": "Mortgage",
    "RELIANCE": "Reliance",
    "UBER": "Uber",
    "TRAVEL": "Travel",
    "CHARITY": "Charity",
    "MISC": "Misc",
    "APPLE SUB": "Apple Sub",
    "SPOTIFY SUB": "Spotify",
    "SPOTIFY": "Spotify",
    "PRIME VIDEO SUB": "Prime Video",
    "NETFLIX": "Netflix",
    "YOUTUBE": "YouTube",
    "CANVA SUB": "Canva Sub",
    "IPSY SUB": "Ipsy Sub",
    "NEWSPAPER": "Newspaper",
    "MUNICIPAL TAXES": "Municipal Taxes",
    "HYDRO": "Hydro",
    "ENBRIDGE": "Gas (Utility)",
}


def normalize_category(raw: str) -> str:
    if not raw:
        return "Misc"
    upper = str(raw).strip().upper()
    return CATEGORY_MAP.get(upper, str(raw).strip().title())


def to_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def detect_format(sheet_names: list[str]) -> str:
    for name in sheet_names:
        if re.match(r"CC - \w+ \d{2}$", name):
            return "fy24"
    for name in sheet_names:
        if "(CC)" in name:
            return "fy25_26"
    return "unknown"


def import_xlsx(file_path: str, db: Session) -> dict:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    fmt = detect_format(wb.sheetnames)

    counts = {"transactions": 0, "income": 0, "skipped": 0}

    if fmt == "fy24":
        _import_fy24(wb, db, counts)
    elif fmt == "fy25_26":
        _import_fy25_26(wb, db, counts)
    else:
        raise ValueError("Unrecognized spreadsheet format")

    db.commit()
    return counts


# ---------------------------------------------------------------------------
# FY24 importer
# ---------------------------------------------------------------------------

def _parse_month_year_fy24(sheet_name: str) -> tuple[int, int] | None:
    """'CC - Jan 24' -> (1, 2024)"""
    m = re.match(r"(?:CC|CB) - (\w+) (\d{2})$", sheet_name)
    if not m:
        return None
    month_str, yr_short = m.group(1).lower(), m.group(2)
    month = MONTH_NAMES.get(month_str)
    year = 2000 + int(yr_short)
    if month:
        return month, year
    return None


def _import_fy24(wb, db: Session, counts: dict):
    # Import income from "2024 Balance Sheet"
    if "2024 Balance Sheet" in wb.sheetnames:
        _import_income_fy24(wb["2024 Balance Sheet"], db, counts)

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("CC - "):
            continue
        result = _parse_month_year_fy24(sheet_name)
        if not result:
            continue
        month, year = result
        _import_fy24_cc_sheet(wb[sheet_name], month, year, db, counts)


def _import_income_fy24(ws, db: Session, counts: dict):
    """Parse '2024 Balance Sheet': Month | Net Income (Matt) | Comm Income (Matt) | Net Income (Nicole)"""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not isinstance(row[0], str):
            continue
        month_key = row[0].strip().lower()
        if month_key not in MONTH_SHORT:  # exact 3-letter match only
            continue
        month = MONTH_SHORT[month_key]
        year = 2024
        _upsert_income(db, year, month, "Matt", "base", row[1] or 0, counts)
        _upsert_income(db, year, month, "Matt", "commission", row[2] or 0, counts)
        _upsert_income(db, year, month, "Nicole", "base", row[3] or 0, counts)


def _import_fy24_cc_sheet(ws, month: int, year: int, db: Session, counts: dict):
    """
    FY24 CC sheet has VISA on left (cols B-E) and AMEX on right (cols E-H).
    Headers row: (None, 'ITEM', 'DATE', 'CHARGE', 'CATEGORY', ...)
    """
    rows = list(ws.iter_rows(values_only=True))

    # Find header row
    header_row = None
    for i, row in enumerate(rows):
        if row[1] == "ITEM" and row[2] == "DATE":
            header_row = i
            break
    if header_row is None:
        return

    for row in rows[header_row + 1:]:
        # VISA: cols 1(B), 2(C), 3(D), 4(E) = ITEM, DATE, CHARGE, CATEGORY
        _add_fy24_txn(row, 1, 2, 3, 4, month, year, "visa", db, counts)
        # AMEX: cols 5(F), 6(G?), check if there's a second card
        # AMEX header was at col 3 (AMEX STATEMENT) in row 0
        # Actual AMEX data cols: 5=ITEM, 6=DATE?, 7=CHARGE, 8=CATEGORY
        # Let's check if col 5 has data
        if len(row) > 8 and row[5] and row[5] != "ITEM":
            _add_fy24_txn(row, 5, 6, 7, 8, month, year, "amex", db, counts)


def _add_fy24_txn(row, item_col, date_col, amt_col, cat_col, month, year, source, db, counts):
    if len(row) <= cat_col:
        return
    merchant = row[item_col]
    txn_date = to_date(row[date_col])
    amount = row[amt_col]
    category = row[cat_col]

    if not merchant or not isinstance(merchant, str):
        return
    if amount is None or not isinstance(amount, (int, float)):
        return
    if amount <= 0:  # skip payments/credits
        return
    if not txn_date:
        return
    if category and str(category).strip().upper() == "PAYMENT":
        return

    txn = Transaction(
        date=txn_date,
        merchant=str(merchant).strip(),
        amount=float(amount),
        category=normalize_category(category),
        year=year,
        month=month,
        source=source,
    )
    db.add(txn)
    counts["transactions"] += 1


# ---------------------------------------------------------------------------
# FY25/FY26 importer
# ---------------------------------------------------------------------------

def _parse_month_year_fy25_26(sheet_name: str) -> tuple[int, int] | None:
    """
    'Jan-2025 (CC)' -> (1, 2025)
    'Jan - 2026 (CC)' -> (1, 2026)
    """
    m = re.match(r"(\w+)\s*[-–]\s*(\d{4})", sheet_name)
    if not m:
        return None
    month_str = m.group(1).strip().lower()[:3]
    year = int(m.group(2))
    month = MONTH_NAMES.get(month_str)
    if month:
        return month, year
    return None


def _import_fy25_26(wb, db: Session, counts: dict):
    # Detect year from CC sheet names (most reliable source)
    detected_year = None
    for name in wb.sheetnames:
        r = _parse_month_year_fy25_26(name)
        if r:
            detected_year = r[1]
            break

    # Import income from Summary sheet
    summary_name = next((s for s in wb.sheetnames if s.lower() == "summary"), None)
    if summary_name:
        _import_income_fy25_26(wb[summary_name], detected_year, db, counts)

    for sheet_name in wb.sheetnames:
        if "(CC)" not in sheet_name and "(cc)" not in sheet_name:
            continue
        result = _parse_month_year_fy25_26(sheet_name)
        if not result:
            continue
        month, year = result
        _import_fy25_26_cc_sheet(wb[sheet_name], month, year, db, counts)


def _import_income_fy25_26(ws, detected_year: int | None, db: Session, counts: dict):
    """
    FY25 Summary: row per month — Month | Net Income (Matt) | Comm Income (Matt) | Net Income (Nicole)
    FY26 Summary: column per month — 'Matt Pay' / 'Nicole Pay' rows with month columns
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return

    # Detect FY26 layout: scan for any row that has a year >= 2024 in col 1 OR 'Matt Pay' row
    is_fy26 = any(
        (isinstance(r[1], (int, float)) and 2020 <= r[1] <= 2035) or
        (isinstance(r[1], str) and "matt pay" in r[1].lower())
        for r in rows if r
    )
    if is_fy26:
        _import_income_fy26(rows, detected_year, db, counts)
        _import_expenses_fy26_summary(rows, detected_year, db, counts)
        return

    # FY25 layout: one row per month with exact 3-letter month key in col 0
    year = detected_year
    if year is None:
        return
    for row in rows:
        if not row[0] or not isinstance(row[0], str):
            continue
        key = row[0].strip().lower()
        if key not in MONTH_SHORT:
            continue
        month = MONTH_SHORT[key]
        _upsert_income(db, year, month, "Matt", "base", row[1] or 0, counts)
        _upsert_income(db, year, month, "Matt", "commission", row[2] or 0, counts)
        _upsert_income(db, year, month, "Nicole", "base", row[3] or 0, counts)


def _import_income_fy26(rows, detected_year: int | None, db: Session, counts: dict):
    """
    FY26 Summary layout:
    One row has: (None, 2026, 'JAN', 'FEB', ..., 'DEC', ...)
    Other rows: (None, 'Matt Pay', val, val, ...) and (None, 'Nicole Pay', val, val, ...)
    """
    col_months = {}
    month_map = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
                 "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}

    # Find the header row with month column labels
    for row in rows:
        has_months = any(
            isinstance(v, str) and v.strip().lower()[:3] in month_map
            for v in (row or [])
        )
        if has_months:
            for col_idx, val in enumerate(row):
                if isinstance(val, str) and val.strip().lower()[:3] in month_map:
                    col_months[col_idx] = month_map[val.strip().lower()[:3]]
            break  # use first matching header row

    year = detected_year
    if year is None:
        # Try to find year from a row that has it
        for row in rows:
            if row and isinstance(row[1], (int, float)) and row[1] >= 2024:
                year = int(row[1])
                break
    if year is None:
        year = 2026

    for row in rows[1:]:
        if not row[1] or not isinstance(row[1], str):
            continue
        label = row[1].strip().lower()
        if "matt" in label and "pay" in label:
            person, itype = "Matt", "base"
        elif "nicole" in label and "pay" in label:
            person, itype = "Nicole", "base"
        else:
            continue
        for col_idx, month in col_months.items():
            val = row[col_idx] if col_idx < len(row) else None
            if isinstance(val, (int, float)) and val > 0:
                _upsert_income(db, year, month, person, itype, float(val), counts)


_SKIP_SUMMARY_LABELS = {
    "", "needs", "wants", "total", "total expenses", "expenses",
}


def _map_summary_expense_category(label: str) -> str:
    l = label.lower()
    if "housing" in l or "mortgage" in l:   return "Mortgage"
    if "enbridge" in l:                      return "Gas (Utility)"
    if "hydro" in l:                         return "Hydro"
    if "groceries" in l or "grocery" in l:  return "Groceries"
    if "pet" in l:                           return "Pets"
    if "transportation" in l:               return "Transportation"
    if "internet" in l:                      return "Internet"
    if "security" in l:                      return "Security"
    if "mobile" in l:                        return "Mobile"
    if "insurance" in l:                     return "Insurance"
    if "property tax" in l:                  return "Municipal Taxes"
    if "debit payment" in l or "loan" in l: return "Debt Payment"
    if "subscription" in l:                  return "Entertainment"
    if "entertainment" in l or "dining" in l: return "Entertainment"
    if "clothing" in l or "personal care" in l: return "Clothes"
    if "gift" in l or "donation" in l:      return "Gifts"
    if "charity" in l:                       return "Charity"
    if "travel" in l:                        return "Travel"
    if "coffee" in l:                        return "Coffee"
    if "home" in l:                          return "Home"
    if "misc" in l or "alcohol" in l or "cannabis" in l: return "Misc"
    return label.strip().title()


def _import_expenses_fy26_summary(rows, detected_year: int | None, db: Session, counts: dict):
    """
    Import per-month expense rows from the FY26 Summary sheet.
    Rows look like: (None, 'Housing', 1967.98, 1967.98, ...)
    with month columns identified by the same header row used for income.
    """
    month_map = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
                 "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}

    # Find col -> month mapping
    col_months: dict[int, int] = {}
    for row in rows:
        has_months = any(
            isinstance(v, str) and v.strip().lower()[:3] in month_map
            for v in (row or [])
        )
        if has_months:
            for col_idx, val in enumerate(row):
                if isinstance(val, str) and val.strip().lower()[:3] in month_map:
                    col_months[col_idx] = month_map[val.strip().lower()[:3]]
            break

    if not col_months:
        return

    year = detected_year or 2026

    for row in rows:
        if not row or not row[1] or not isinstance(row[1], str):
            continue
        label = row[1].strip()
        if label.lower() in _SKIP_SUMMARY_LABELS:
            continue
        lower = label.lower()
        # Skip income rows already handled
        if ("matt" in lower and "pay" in lower) or ("nicole" in lower and "pay" in lower):
            continue

        category = _map_summary_expense_category(label)

        for col_idx, month in col_months.items():
            val = row[col_idx] if col_idx < len(row) else None
            if not isinstance(val, (int, float)) or val <= 0:
                continue
            txn = Transaction(
                date=date(year, month, 1),
                merchant=label,
                amount=float(val),
                category=category,
                year=year,
                month=month,
                source="summary",
            )
            db.add(txn)
            counts["transactions"] += 1


def _import_fy25_26_cc_sheet(ws, month: int, year: int, db: Session, counts: dict):
    """
    Headers at row 1: DATE, ITEM, AMOUNT, CATEGORY
    """
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if not row or len(row) < 4:
            continue
        txn_date = to_date(row[0])
        merchant = row[1]
        amount = row[2]
        category = row[3]

        if not merchant or not isinstance(merchant, str):
            continue
        if amount is None or not isinstance(amount, (int, float)):
            continue
        if amount <= 0:
            continue
        if not txn_date:
            continue
        if category and str(category).strip().upper() == "PAYMENT":
            continue

        txn = Transaction(
            date=txn_date,
            merchant=str(merchant).strip(),
            amount=float(amount),
            category=normalize_category(category),
            year=year,
            month=month,
            source="cc",
        )
        db.add(txn)
        counts["transactions"] += 1


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _upsert_income(db: Session, year: int, month: int, person: str, itype: str, amount: float, counts: dict):
    from models import Income
    from sqlalchemy import select
    existing = db.execute(
        select(Income).where(
            Income.year == year,
            Income.month == month,
            Income.person == person,
            Income.income_type == itype,
        )
    ).scalar_one_or_none()
    if existing:
        existing.amount = amount
        counts["skipped"] += 1
    else:
        db.add(Income(year=year, month=month, person=person, income_type=itype, amount=amount))
        counts["income"] += 1
